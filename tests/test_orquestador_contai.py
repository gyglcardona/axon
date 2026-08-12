"""
Pruebas de la exportación a Contai en orquestador.py: configuración de
cuentas, import de plan de cuentas/terceros, y previsualizar/confirmar la
exportación (arma los .xlsx en memoria, marca estado_contai). Nunca toca
Excel real del disco del usuario -- los .xlsx de prueba se arman con
openpyxl en tmp_path, mismo layout confirmado contra los archivos reales
que compartió el usuario.
"""

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import orquestador  # noqa: E402
import state_store  # noqa: E402
from dian_parser import FacturaDian  # noqa: E402
from motor_reglas import ItemSiigo, ResultadoClasificacion  # noqa: E402


@pytest.fixture
def empresa_configurada(tmp_path, monkeypatch):
    registro = tmp_path / "registro.json"
    registro.write_text(
        '{"empresas":[{"slug":"empresa-test","nit":"900000000","nombre":"EMPRESA TEST"}]}',
        encoding="utf-8",
    )
    monkeypatch.setattr(orquestador, "REGISTRO", registro)
    monkeypatch.setattr(orquestador, "CONFIG_EMPRESAS_DIR", tmp_path / "config" / "empresas")
    monkeypatch.setattr(orquestador, "BASE_DATOS_EMPRESAS", tmp_path / "data" / "empresas")

    original_conectar = state_store.conectar

    def _conectar_en_tmp(nit_empresa, base_dir=None):
        return original_conectar(nit_empresa, base_dir=tmp_path / "data" / "empresas")

    monkeypatch.setattr(state_store, "conectar", _conectar_en_tmp)
    return "empresa-test", tmp_path


def _sembrar_factura(cufe, proveedor_nit="900111222", con_iva=True, cuenta_contable="620505"):
    conn = state_store.conectar("900000000")
    impuestos = [{"tipo": "IVA", "porcentaje": 19.0, "valor": 19000.0}] if con_iva else []
    factura = FacturaDian(
        cufe=cufe, numero_factura="FE1102", prefijo="FE", numero_puro="1102", fecha_emision="2025-01-31",
        proveedor_nombre="PROVEEDOR TEST", proveedor_nit=proveedor_nit,
        proveedor_correo=None, proveedor_direccion=None,
        subtotal_xml=100000, subtotal_fuente="TaxExclusiveAmount", total_pagar_xml=119000,
    )
    items = [ItemSiigo(descripcion="ITEM", cantidad=1, valor_unitario=100000, cuenta_contable=cuenta_contable,
                        impuestos=impuestos)]
    resultado = ResultadoClasificacion(factura=factura, items=items, resuelto_por="manual")
    state_store.guardar_resultado(conn, resultado, archivo_origen=Path("x.zip"))
    conn.close()


def _config_lista(slug):
    orquestador.guardar_config_contai(slug, {
        "cuenta_credito_contado": "110505",
        "cuentas_iva_por_tarifa": {"19.0": "24081001"},
        "cuentas_gasto_por_categoria": {"19.0": "620505", "0.0": "620505"},
    })


def _crear_xlsx_plan_cuentas_contai(ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Código Cuenta", "Nombre Cuenta", "Tipo de Cuenta", "Id. Recibe Movto.",
               "Id. Centro Costo", "Id. Ajustes", "Porcentaje Base", "Tipo Plazo", "Activo"])
    ws.append(["1", "ACTIVO", "N", "N", "N", "N", 0, "N", "S"])
    ws.append(["110505", "CAJA GENERAL", "N", "S", "N", "N", 0, "N", "S"])
    wb.save(ruta)


def _crear_xlsx_terceros_contai(ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    columnas = ["NIT", "Tipo", "Nombre", "Direccion", "Ciudad", "Telefono", "Municipio",
                "Activo", "Tiene RUT", "Pais", "Email", "Celular", "Plazo",
                "Actividad Económica", "Indicativo", "Naturaleza"]
    ws.append(columnas)
    ws.append(["          A", "T", "Aombre", "", "", "", "", "A", "N", "", "", "", "  0", "", "", ""])  # fila plantilla
    ws.append(["900111222", "P", "PROVEEDOR TEST", "", "", "", "", "S", "N", "", "", "", "0", "", "", "J"])
    wb.save(ruta)


def _crear_xlsx_terceros_contai_con_nit_duplicado(ruta):
    """Caso real (contai_terceros.xlsx): "18576831" y "E18576831" colapsan
    al mismo NIT tras limpiar caracteres no numéricos -- una cédula de
    extranjería exportada con y sin el prefijo "E"."""
    wb = openpyxl.Workbook()
    ws = wb.active
    columnas = ["NIT", "Tipo", "Nombre", "Direccion", "Ciudad", "Telefono", "Municipio",
                "Activo", "Tiene RUT", "Pais", "Email", "Celular", "Plazo",
                "Actividad Económica", "Indicativo", "Naturaleza"]
    ws.append(columnas)
    ws.append(["18576831", "T", "PERSONA UNO", "", "", "", "", "S", "N", "", "", "", "0", "", "", "N"])
    ws.append(["E18576831", "T", "PERSONA UNO", "", "", "", "", "S", "N", "", "", "", "0", "", "", "N"])
    wb.save(ruta)


def _crear_xlsx_comprobantes_contai(ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Comprobante", "Nombre", "Maneja Consecutivo", "Nro. Consecutivo", "Maneja Codificación",
               "Consecutivo Obligatorio", "Vigencia", "Documento No Obligado", "Resolución", "Prefijo"])
    ws.append(["00010", "CAUSACIONES", "N", "000000000", "N", "N", "S", "N", "", ""])
    ws.append(["00003", "FACTURACION ELECTRONICA", "N", "000000000", "N", "S", "S", "N", "", ""])
    wb.save(ruta)


def _crear_xlsx_movimientos_contai(ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Cuenta", "Comprobante", "Fecha(mm/dd/yyyy)", "Documento", "Documento Ref.", "NIT", "Detalle",
               "Tipo", "Valor", "Base", "Centro de Costo", "Trans. Ext", "Plazo", "Docto Electrónico"])
    # mismo proveedor, dos documentos, la cuenta 620505 se repite más que la 620510
    ws.append(["620505", "00010", "01/31/2025", "F1", "F1", "900111222", "COMPRA ENERO 2025", 1, 100000, 0, "", "", 0, ""])
    ws.append(["24081001", "00010", "01/31/2025", "F1", "F1", "900111222", "COMPRA ENERO 2025", 1, 19000, 100000, "", "", 0, ""])
    ws.append(["110505", "00010", "01/31/2025", "F1", "F1", "", "COMPRA ENERO 2025", 2, 119000, 0, "", "", 0, ""])
    ws.append(["620505", "00010", "02/28/2025", "F2", "F2", "900111222", "COMPRA FEBRERO 2025", 1, 50000, 0, "", "", 0, ""])
    ws.append(["620510", "00010", "02/28/2025", "F2", "F2", "900111222", "COMPRA FEBRERO 2025", 1, 20000, 0, "", "", 0, ""])
    ws.append(["110505", "00010", "02/28/2025", "F2", "F2", "", "COMPRA FEBRERO 2025", 2, 70000, 0, "", "", 0, ""])
    wb.save(ruta)


def test_config_contai_default_sin_archivo(empresa_configurada):
    slug, _ = empresa_configurada
    assert orquestador.obtener_config_contai(slug) == orquestador._CONFIG_CONTAI_DEFAULT


def test_guardar_config_contai_hace_merge_no_reemplaza(empresa_configurada):
    slug, _ = empresa_configurada
    orquestador.guardar_config_contai(slug, {"cuenta_credito_contado": "110505"})
    orquestador.guardar_config_contai(slug, {"cuentas_iva_por_tarifa": {"19.0": "24081001"}})

    config = orquestador.obtener_config_contai(slug)
    assert config["cuenta_credito_contado"] == "110505"
    assert config["cuentas_iva_por_tarifa"] == {"19.0": "24081001"}
    assert config["comprobante"] == "00010"  # el default no se pierde


def test_importar_plan_cuentas_contai(empresa_configurada, tmp_path):
    slug, _ = empresa_configurada
    ruta = tmp_path / "plancuentas.xlsx"
    _crear_xlsx_plan_cuentas_contai(ruta)

    resumen = orquestador.importar_plan_cuentas_contai(slug, str(ruta))

    assert resumen == {"total": 2, "transaccionales": 1}
    cuentas = orquestador.listar_plan_cuentas_contai(slug, solo_transaccionales=True)
    assert [c["codigo"] for c in cuentas] == ["110505"]
    # tabla propia de Contai -- no debe tocar el plan de cuentas de Siigo
    assert orquestador.listar_plan_cuentas(slug) == []


def test_importar_terceros_contai_descarta_fila_plantilla(empresa_configurada, tmp_path):
    slug, _ = empresa_configurada
    ruta = tmp_path / "terceros.xlsx"
    _crear_xlsx_terceros_contai(ruta)

    resumen = orquestador.importar_terceros_contai(slug, str(ruta))

    assert resumen == {"total": 1}  # la fila plantilla (NIT no numérico) se descarta
    empresa = orquestador.resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    assert state_store.existe_tercero_contai(conn, "900111222")
    conn.close()


def test_importar_terceros_contai_no_revienta_con_nit_duplicado(empresa_configurada, tmp_path):
    """Dos filas del Excel real colapsan al mismo NIT tras limpiar
    caracteres no numéricos (ver _crear_xlsx_terceros_contai_con_nit_duplicado)
    -- no debe reventar con UNIQUE constraint, la última fila gana."""
    slug, _ = empresa_configurada
    ruta = tmp_path / "terceros_dup.xlsx"
    _crear_xlsx_terceros_contai_con_nit_duplicado(ruta)

    resumen = orquestador.importar_terceros_contai(slug, str(ruta))

    assert resumen == {"total": 1}
    empresa = orquestador.resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    assert state_store.existe_tercero_contai(conn, "18576831")
    assert state_store.contar_terceros_contai(conn) == 1
    conn.close()


def test_importar_comprobantes_contai(empresa_configurada, tmp_path):
    slug, _ = empresa_configurada
    ruta = tmp_path / "comprobantes.xlsx"
    _crear_xlsx_comprobantes_contai(ruta)

    resumen = orquestador.importar_comprobantes_contai(slug, str(ruta))

    assert resumen == {"total": 2}
    comprobantes = orquestador.listar_comprobantes_contai(slug)
    assert {c["Comprobante"] for c in comprobantes} == {"00010", "00003"}
    assert next(c["Nombre"] for c in comprobantes if c["Comprobante"] == "00010") == "CAUSACIONES"


def test_importar_movimientos_contai(empresa_configurada, tmp_path):
    slug, _ = empresa_configurada
    ruta = tmp_path / "movimientos.xlsx"
    _crear_xlsx_movimientos_contai(ruta)

    resumen = orquestador.importar_movimientos_contai(slug, str(ruta))

    assert resumen == {
        "total": 6, "lineas_insertadas": 6, "lineas_omitidas": 0, "documentos_nuevos": 2,
        "total_lineas": 6, "proveedores_distintos": 1,
    }
    assert orquestador.obtener_resumen_movimientos_contai(slug) == {"total_lineas": 6, "proveedores_distintos": 1}


def test_reimportar_el_mismo_archivo_no_duplica_nada(empresa_configurada, tmp_path):
    """Caso real: el usuario reimporta por error el mismo mes -- los
    documentos F1 y F2 ya existen, así que sus líneas se saltan completas,
    no se duplican."""
    slug, _ = empresa_configurada
    ruta = tmp_path / "movimientos.xlsx"
    _crear_xlsx_movimientos_contai(ruta)
    orquestador.importar_movimientos_contai(slug, str(ruta))

    resumen = orquestador.importar_movimientos_contai(slug, str(ruta))

    assert resumen["lineas_insertadas"] == 0
    assert resumen["lineas_omitidas"] == 6
    assert resumen["total_lineas"] == 6  # no se duplicó nada


def test_importar_un_mes_nuevo_se_agrega_al_ya_importado(empresa_configurada, tmp_path):
    """Importar marzo después de ya haber importado enero/febrero (F1, F2)
    debe AGREGAR el mes nuevo, no reemplazar lo que ya había."""
    slug, _ = empresa_configurada
    ruta_ene_feb = tmp_path / "movimientos.xlsx"
    _crear_xlsx_movimientos_contai(ruta_ene_feb)
    orquestador.importar_movimientos_contai(slug, str(ruta_ene_feb))

    ruta_marzo = tmp_path / "movimientos_marzo.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Cuenta", "Comprobante", "Fecha(mm/dd/yyyy)", "Documento", "Documento Ref.", "NIT", "Detalle",
               "Tipo", "Valor", "Base", "Centro de Costo", "Trans. Ext", "Plazo", "Docto Electrónico"])
    ws.append(["620505", "00010", "03/15/2025", "F3", "F3", "900111222", "COMPRA MARZO 2025", 1, 30000, 0, "", "", 0, ""])
    ws.append(["110505", "00010", "03/15/2025", "F3", "F3", "", "COMPRA MARZO 2025", 2, 30000, 0, "", "", 0, ""])
    wb.save(ruta_marzo)

    resumen = orquestador.importar_movimientos_contai(slug, str(ruta_marzo))

    assert resumen["lineas_insertadas"] == 2
    assert resumen["lineas_omitidas"] == 0
    assert resumen["total_lineas"] == 8  # 6 de antes + 2 nuevas
    assert {d["documento"] for d in orquestador.listar_movimientos_contai(slug)} == {"F1", "F2", "F3"}


# --- consultar el histórico ya importado (no solo importarlo) ---


def test_listar_movimientos_contai_agrupa_por_documento(empresa_configurada, tmp_path):
    """El fixture trae 2 documentos (F1, F2) del mismo proveedor -- deben
    verse como 2 filas agrupadas, cada una con sus propias líneas de
    cuenta/tipo/valor, no como 6 líneas sueltas."""
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug)

    assert {d["documento"] for d in documentos} == {"F1", "F2"}
    f1 = next(d for d in documentos if d["documento"] == "F1")
    assert f1["proveedor_nit"] == "900111222"
    assert len(f1["lineas"]) == 3  # 620505 débito, 24081001 débito, 110505 crédito
    assert f1["total_debito"] == 119000  # 100000 + 19000, la línea de crédito (110505) no cuenta
    assert {(l["cuenta"], l["tipo"], l["valor"]) for l in f1["lineas"]} == {
        ("620505", 1, 100000), ("24081001", 1, 19000), ("110505", 2, 119000),
    }


def test_listar_movimientos_contai_resuelve_nombre_via_terceros(empresa_configurada, tmp_path):
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))
    orquestador.importar_terceros_contai(slug, str(_ruta_terceros(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug)

    assert all(d["proveedor_nombre"] == "PROVEEDOR TEST" for d in documentos)


def test_listar_movimientos_contai_sin_terceros_importados_nombre_es_none(empresa_configurada, tmp_path):
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug)

    assert all(d["proveedor_nombre"] is None for d in documentos)


def test_listar_movimientos_contai_filtra_por_numero_de_documento(empresa_configurada, tmp_path):
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug, texto="F2")

    assert [d["documento"] for d in documentos] == ["F2"]


def test_listar_movimientos_contai_filtra_por_nit(empresa_configurada, tmp_path):
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug, texto="900111222")

    assert {d["documento"] for d in documentos} == {"F1", "F2"}


def test_listar_movimientos_contai_filtra_por_rango_de_fechas(empresa_configurada, tmp_path):
    """F1 es 01/31/2025, F2 es 02/28/2025 (ver _crear_xlsx_movimientos_contai) --
    un rango que solo cubre enero debe traer solo F1."""
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug, desde="2025-01-01", hasta="2025-01-31")

    assert [d["documento"] for d in documentos] == ["F1"]


def test_listar_movimientos_contai_rango_de_fechas_fuera_de_rango_da_vacio(empresa_configurada, tmp_path):
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug, desde="2025-06-01", hasta="2025-06-30")

    assert documentos == []


def test_listar_movimientos_contai_solo_desde_sin_hasta(empresa_configurada, tmp_path):
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug, desde="2025-02-01")

    assert [d["documento"] for d in documentos] == ["F2"]


def test_listar_movimientos_contai_filtra_por_nombre_de_proveedor(empresa_configurada, tmp_path):
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))
    orquestador.importar_terceros_contai(slug, str(_ruta_terceros(tmp)))

    documentos = orquestador.listar_movimientos_contai(slug, texto="proveedor test")  # minúsculas, parcial

    assert {d["documento"] for d in documentos} == {"F1", "F2"}


def test_listar_movimientos_contai_sin_coincidencias_da_lista_vacia(empresa_configurada, tmp_path):
    slug, tmp = empresa_configurada
    orquestador.importar_movimientos_contai(slug, str(_ruta_movimientos(tmp)))

    assert orquestador.listar_movimientos_contai(slug, texto="no-existe") == []


def test_listar_movimientos_contai_sin_importar_nada_da_lista_vacia(empresa_configurada):
    slug, _ = empresa_configurada
    assert orquestador.listar_movimientos_contai(slug) == []


def _ruta_movimientos(tmp_path):
    ruta = tmp_path / "movimientos.xlsx"
    _crear_xlsx_movimientos_contai(ruta)
    return ruta


def _ruta_terceros(tmp_path):
    ruta = tmp_path / "terceros.xlsx"
    _crear_xlsx_terceros_contai(ruta)
    return ruta


def test_sugerencia_historico_contai_por_proveedor(empresa_configurada, tmp_path):
    """Con destino_causacion=contai y el histórico de movimientos importado,
    una línea de gasto sin cuenta debe recibir la cuenta más frecuente de
    ese proveedor en el histórico (620505, aparece en los dos documentos del
    fixture) -- nunca una cuenta de IVA (24081001), aunque también sea
    débito y también aparezca para ese proveedor."""
    slug, tmp = empresa_configurada
    orquestador.guardar_destino_causacion(slug, "contai")
    orquestador.guardar_config_contai(slug, {"cuentas_iva_por_tarifa": {"19.0": "24081001"}})
    ruta = tmp / "movimientos.xlsx"
    _crear_xlsx_movimientos_contai(ruta)
    orquestador.importar_movimientos_contai(slug, str(ruta))

    conn = state_store.conectar("900000000")
    factura = FacturaDian(
        cufe="CUFE-HIST", numero_factura="FE9", prefijo="FE", numero_puro="9", fecha_emision="2025-03-31",
        proveedor_nombre="PROVEEDOR TEST", proveedor_nit="900111222",
        proveedor_correo=None, proveedor_direccion=None,
        subtotal_xml=100000, subtotal_fuente="TaxExclusiveAmount", total_pagar_xml=119000,
    )
    item = ItemSiigo(descripcion="ITEM SIN CUENTA", cantidad=1, valor_unitario=100000, cuenta_contable=None,
                      impuestos=[], origen="xml")
    resultado = ResultadoClasificacion(factura=factura, items=[item], resuelto_por="manual")

    orquestador._aplicar_sugerencias(conn, resultado, slug=slug)
    conn.close()

    assert item.cuenta_contable == "620505"
    assert resultado.resuelto_por == "historico"


def test_sugerencia_historico_contai_no_aplica_a_empresa_siigo(empresa_configurada, tmp_path):
    """La misma sugerencia por historial de Contai no debe activarse para
    una empresa con destino_causacion=siigo (default), aunque por alguna
    razón tuviera datos en movimientos_contai_historico."""
    slug, tmp = empresa_configurada
    ruta = tmp / "movimientos.xlsx"
    _crear_xlsx_movimientos_contai(ruta)
    orquestador.importar_movimientos_contai(slug, str(ruta))

    conn = state_store.conectar("900000000")
    factura = FacturaDian(
        cufe="CUFE-HIST2", numero_factura="FE10", prefijo="FE", numero_puro="10", fecha_emision="2025-03-31",
        proveedor_nombre="PROVEEDOR TEST", proveedor_nit="900111222",
        proveedor_correo=None, proveedor_direccion=None,
        subtotal_xml=100000, subtotal_fuente="TaxExclusiveAmount", total_pagar_xml=119000,
    )
    item = ItemSiigo(descripcion="ITEM SIN CUENTA", cantidad=1, valor_unitario=100000, cuenta_contable=None,
                      impuestos=[], origen="xml")
    resultado = ResultadoClasificacion(factura=factura, items=[item], resuelto_por="manual")

    orquestador._aplicar_sugerencias(conn, resultado, slug=slug)
    conn.close()

    assert item.cuenta_contable is None


def test_previsualizar_exportacion_marca_tercero_nuevo(empresa_configurada):
    slug, _ = empresa_configurada
    _config_lista(slug)
    _sembrar_factura("CUFE-1")

    r = orquestador.previsualizar_exportacion_contai(slug, ["CUFE-1"])

    assert r[0]["exportable"] is True
    assert r[0]["tercero_nuevo"] is True  # todavía no se importó ningún maestro de terceros
    assert len(r[0]["filas"]) == 3  # 1 gasto + 1 IVA + 1 crédito a caja


def test_previsualizar_exportacion_respeta_modo_de_pago_de_la_factura(empresa_configurada):
    """La empresa tiene modo_pago_default='contado' (110505), pero esta
    factura puntual se marcó a mano como 'credito' (220501) -- la
    previsualización debe usar la cuenta de crédito de la factura."""
    slug, _ = empresa_configurada
    orquestador.guardar_config_contai(slug, {
        "cuenta_credito_contado": "110505",
        "cuenta_credito_credito": "220501",
        "cuentas_iva_por_tarifa": {"19.0": "24081001"},
        "cuentas_gasto_por_categoria": {"19.0": "620505", "0.0": "620505"},
    })
    _sembrar_factura("CUFE-1")
    orquestador.actualizar_factura(slug, "CUFE-1", {"modo_pago_contai": "credito"})

    r = orquestador.previsualizar_exportacion_contai(slug, ["CUFE-1"])

    assert r[0]["exportable"] is True
    credito = next(f for f in r[0]["filas"] if f["Tipo"] == 2)
    assert credito["Cuenta"] == "220501"


def test_previsualizar_exportacion_bloqueada_no_trae_filas(empresa_configurada):
    slug, _ = empresa_configurada
    _sembrar_factura("CUFE-1")  # sin _config_lista -- no hay cuentas configuradas

    r = orquestador.previsualizar_exportacion_contai(slug, ["CUFE-1"])

    assert r[0]["exportable"] is False
    assert r[0]["filas"] is None
    assert r[0]["motivos_bloqueo"]


def test_confirmar_exportacion_genera_xlsx_valido_y_marca_estado(empresa_configurada):
    slug, _ = empresa_configurada
    _config_lista(slug)
    _sembrar_factura("CUFE-1")

    resultado = orquestador.confirmar_exportacion_contai(slug, ["CUFE-1"])

    assert resultado["exportadas"] == 1
    assert resultado["con_error"] == 0

    wb = openpyxl.load_workbook(io.BytesIO(resultado["movimientos_xlsx"]))
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    assert filas[0] == tuple(__import__("contai_export").COLUMNAS_MOVIMIENTO)
    assert len(filas) == 1 + 3  # encabezado + 1 gasto + 1 IVA + 1 crédito

    # el .txt trae las mismas filas, separadas por ';' (confirmado con el usuario)
    lineas_txt = resultado["movimientos_txt"].decode("utf-8-sig").strip("\r\n").split("\r\n")
    assert lineas_txt[0] == ";".join(__import__("contai_export").COLUMNAS_MOVIMIENTO)
    assert len(lineas_txt) == 1 + 3

    facturas = orquestador.listar_facturas(slug)
    assert facturas[0]["estado_contai"] == "exportado"


def test_confirmar_exportacion_bloqueada_no_cambia_estado(empresa_configurada):
    slug, _ = empresa_configurada
    _sembrar_factura("CUFE-1")  # sin config -- queda bloqueada

    resultado = orquestador.confirmar_exportacion_contai(slug, ["CUFE-1"])

    assert resultado["exportadas"] == 0
    assert resultado["con_error"] == 1
    facturas = orquestador.listar_facturas(slug)
    assert facturas[0]["estado_contai"] == "pendiente"


def test_confirmar_exportacion_siempre_genera_terceros_de_lo_exportado(empresa_configurada, monkeypatch, tmp_path):
    """El archivo de terceros ya no depende de si el proveedor es "nuevo"
    para el maestro de Contai -- se genera siempre, con todos los
    proveedores de las facturas que sí quedaron exportadas (pedido
    explícito del usuario: necesita el lote completo, no solo lo nuevo)."""
    slug, _ = empresa_configurada
    _config_lista(slug)
    _sembrar_factura("CUFE-1", proveedor_nit="900111222")

    monkeypatch.setattr(orquestador, "_extraer_tercero_de_origen", lambda archivo_origen, cufe: {
        "nombre": "PROVEEDOR TEST", "nit": "900111222", "id_type": "31",
        "direccion": None, "correo": None, "telefono": None,
    })

    resultado = orquestador.confirmar_exportacion_contai(slug, ["CUFE-1"])

    assert resultado["terceros_xlsx"] is not None
    wb = openpyxl.load_workbook(io.BytesIO(resultado["terceros_xlsx"]))
    filas = list(wb.active.iter_rows(values_only=True))
    assert filas[1][0] == "900111222"  # NIT en la primera fila de datos

    lineas_txt = resultado["terceros_txt"].decode("utf-8-sig").strip("\r\n").split("\r\n")
    assert lineas_txt[1].split(";")[0] == "900111222"


def test_confirmar_exportacion_incluye_tercero_aunque_ya_exista_en_el_maestro(empresa_configurada, monkeypatch, tmp_path):
    """Antes, si el proveedor ya estaba en terceros_contai (importado de
    Contai), se lo saltaba -- ahora igual debe aparecer en el archivo,
    porque el criterio ya no es "es nuevo", es "está en el lote exportado"."""
    slug, _ = empresa_configurada
    _config_lista(slug)
    ruta = tmp_path / "terceros.xlsx"
    _crear_xlsx_terceros_contai(ruta)
    orquestador.importar_terceros_contai(slug, str(ruta))
    _sembrar_factura("CUFE-1", proveedor_nit="900111222")

    monkeypatch.setattr(orquestador, "_extraer_tercero_de_origen", lambda archivo_origen, cufe: {
        "nombre": "PROVEEDOR TEST", "nit": "900111222", "id_type": "31",
        "direccion": None, "correo": None, "telefono": None,
    })

    resultado = orquestador.confirmar_exportacion_contai(slug, ["CUFE-1"])

    assert resultado["terceros_xlsx"] is not None
    wb = openpyxl.load_workbook(io.BytesIO(resultado["terceros_xlsx"]))
    filas = list(wb.active.iter_rows(values_only=True))
    assert filas[1][0] == "900111222"


def test_confirmar_exportacion_sin_facturas_exportables_genera_terceros_vacio(empresa_configurada):
    slug, _ = empresa_configurada
    _sembrar_factura("CUFE-1")  # sin config -- queda bloqueada, nada se exporta

    resultado = orquestador.confirmar_exportacion_contai(slug, ["CUFE-1"])

    assert resultado["terceros_xlsx"] is not None
    wb = openpyxl.load_workbook(io.BytesIO(resultado["terceros_xlsx"]))
    filas = list(wb.active.iter_rows(values_only=True))
    assert len(filas) == 1  # solo el encabezado, ningún proveedor

    lineas_txt = resultado["terceros_txt"].decode("utf-8-sig").strip("\r\n").split("\r\n")
    assert len(lineas_txt) == 1  # solo el encabezado
