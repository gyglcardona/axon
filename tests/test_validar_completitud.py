"""
Pruebas del validador de completitud (listado DIAN vs. bandeja ya importada).
El formato del listado está confirmado contra un archivo real (ver
docs/03-ingesta-dian/validador-completitud.md y src/orquestador.py) --
incluido el detalle de que esa empresa real trae el encabezado "Fecha
Emisión" con el carácter de reemplazo U+FFFD en vez de "ó" (problema del
propio exportador de la DIAN). Acá se arma un .xlsx sintético en tmp_path
para no depender de datos reales de ninguna empresa.
"""

import io
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import orquestador  # noqa: E402
import state_store  # noqa: E402
from dian_parser import FacturaDian  # noqa: E402
from motor_reglas import ItemSiigo, ResultadoClasificacion  # noqa: E402

ENCABEZADOS = [
    "Tipo de documento", "CUFE/CUDE", "Folio", "Prefijo", "Divisa", "Forma de Pago", "Medio de Pago",
    "Fecha Emisión", "Fecha Recepción", "NIT Emisor", "Nombre Emisor", "NIT Receptor", "Nombre Receptor",
    "IVA", "ICA", "IC", "INC", "Timbre", "INC Bolsas", "IN Carbono", "IN Combustibles", "IC Datos",
    "ICL", "INPP", "IBUA", "ICUI", "Rete IVA", "Rete Renta", "Rete ICA", "Total", "Estado", "Grupo",
]


def _fila(cufe, folio, nit_receptor, fecha="01-07-2026", tipo="Factura electrónica", grupo="Recibido",
          nit_emisor="900111222", nombre_emisor="PROVEEDOR TEST", total=100000, encabezado_fecha_corrupto=False):
    return [
        tipo, cufe, folio, "FE", "COP", "1", "10", fecha, fecha, nit_emisor, nombre_emisor,
        nit_receptor, "EMPRESA TEST",
        0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,  # IVA..Rete ICA (16 columnas)
        total, "Aprobado", grupo,
    ]


def _crear_listado(ruta: Path, filas: list, encabezado_fecha_corrupto: bool = False):
    wb = openpyxl.Workbook()
    ws = wb.active
    encabezados = list(ENCABEZADOS)
    if encabezado_fecha_corrupto:
        encabezados[7] = "Fecha Emisi�n"  # tal como llega en el archivo real
    ws.append(encabezados)
    for fila in filas:
        ws.append(fila)
    wb.save(ruta)


@pytest.fixture
def empresa_configurada(tmp_path, monkeypatch):
    registro = tmp_path / "registro.json"
    registro.write_text(
        '{"empresas":[{"slug":"empresa-test","nit":"900000000","nombre":"EMPRESA TEST"}]}',
        encoding="utf-8",
    )
    monkeypatch.setattr(orquestador, "REGISTRO", registro)
    monkeypatch.setattr(orquestador, "CONFIG_EMPRESAS_DIR", tmp_path / "config" / "empresas")
    monkeypatch.setattr(orquestador, "ENTRADA_DIAN", tmp_path / "data" / "entrada-dian")
    monkeypatch.setattr(orquestador, "BASE_DATOS_EMPRESAS", tmp_path / "data" / "empresas")

    original_conectar = state_store.conectar

    def _conectar_en_tmp(nit_empresa, base_dir=None):
        return original_conectar(nit_empresa, base_dir=tmp_path / "data" / "empresas")

    monkeypatch.setattr(state_store, "conectar", _conectar_en_tmp)
    return "empresa-test", tmp_path


def _sembrar_factura(cufe, fecha_emision, nit="900000000"):
    conn = state_store.conectar(nit)
    factura = FacturaDian(
        cufe=cufe, numero_factura="F1", prefijo="F", numero_puro="1", fecha_emision=fecha_emision,
        proveedor_nombre="PROVEEDOR TEST", proveedor_nit="900111222",
        proveedor_correo=None, proveedor_direccion=None,
        subtotal_xml=100000, subtotal_fuente="TaxExclusiveAmount", total_pagar_xml=119000,
    )
    resultado = ResultadoClasificacion(
        factura=factura,
        items=[ItemSiigo(descripcion="ITEM", cantidad=1, valor_unitario=100000, cuenta_contable=None)],
        resuelto_por="manual",
    )
    state_store.guardar_resultado(conn, resultado, archivo_origen=Path("x.zip"))
    conn.close()


def test_sin_rango_de_fechas_da_error_claro(empresa_configurada):
    slug, tmp_path = empresa_configurada
    ruta_carpeta = tmp_path / "data" / "entrada-dian" / slug / "2026" / "07"
    ruta_carpeta.mkdir(parents=True)
    _crear_listado(ruta_carpeta / "listado.xlsx", [_fila("CUFE-1", "1", "900000000")])

    with pytest.raises(ValueError, match="nunca se corre contra toda la tabla"):
        orquestador.validar_completitud(slug, "2026/07", "listado.xlsx", "", "")


def test_archivo_inexistente_da_error_claro(empresa_configurada):
    slug, _ = empresa_configurada
    with pytest.raises(FileNotFoundError):
        orquestador.validar_completitud(slug, "2026/07", "no-existe.xlsx", "2026-07-01", "2026-07-31")


def test_formato_invalido_sin_columnas_esperadas_da_error(empresa_configurada):
    slug, tmp_path = empresa_configurada
    ruta_carpeta = tmp_path / "data" / "entrada-dian" / slug / "2026" / "07"
    ruta_carpeta.mkdir(parents=True)
    wb = openpyxl.Workbook()
    wb.active.append(["Columna A", "Columna B"])
    wb.active.append(["x", "y"])
    wb.save(ruta_carpeta / "raro.xlsx")

    with pytest.raises(ValueError, match="no tiene el formato esperado"):
        orquestador.validar_completitud(slug, "2026/07", "raro.xlsx", "2026-07-01", "2026-07-31")


def test_filtra_por_grupo_nit_tipo_y_encuentra_faltantes(empresa_configurada):
    slug, tmp_path = empresa_configurada
    ruta_carpeta = tmp_path / "data" / "entrada-dian" / slug / "2026" / "07"
    ruta_carpeta.mkdir(parents=True)
    filas = [
        _fila("CUFE-YA-IMPORTADA", "1", "900000000"),
        _fila("CUFE-FALTANTE", "2", "900000000"),
        _fila("CUFE-OTRA-EMPRESA", "3", "800000000"),  # NIT Receptor distinto -- se descarta
        _fila("CUFE-EMITIDO", "4", "900000000", grupo="Emitido"),  # no es compra -- se descarta
        _fila("CUFE-NOTA-CREDITO", "5", "900000000", tipo="Nota crédito"),  # no es factura -- se descarta
    ]
    _crear_listado(ruta_carpeta / "listado.xlsx", filas)
    _sembrar_factura("CUFE-YA-IMPORTADA", "2026-07-01")

    r = orquestador.validar_completitud(slug, "2026/07", "listado.xlsx", "2026-07-01", "2026-07-31")

    assert r["total_listado"] == 2  # solo las 2 que sí son "Recibido" + NIT correcto + "Factura electrónica"
    assert r["total_bandeja_en_rango"] == 1
    assert [f["cufe"] for f in r["faltantes"]] == ["CUFE-FALTANTE"]
    assert r["sobrantes_en_bandeja"] == 0


def test_respeta_el_rango_de_fechas_no_compara_fuera_de_el(empresa_configurada):
    slug, tmp_path = empresa_configurada
    ruta_carpeta = tmp_path / "data" / "entrada-dian" / slug / "2026" / "07"
    ruta_carpeta.mkdir(parents=True)
    filas = [
        _fila("CUFE-JUNIO", "1", "900000000", fecha="15-06-2026"),  # fuera de rango -- se descarta
        _fila("CUFE-JULIO", "2", "900000000", fecha="15-07-2026"),
    ]
    _crear_listado(ruta_carpeta / "listado.xlsx", filas)
    # aunque exista una factura de junio ya importada, el rango de julio no debe verla ni compararla
    _sembrar_factura("CUFE-JUNIO", "2026-06-15")

    r = orquestador.validar_completitud(slug, "2026/07", "listado.xlsx", "2026-07-01", "2026-07-31")

    assert r["total_listado"] == 1
    assert r["total_bandeja_en_rango"] == 0
    assert [f["cufe"] for f in r["faltantes"]] == ["CUFE-JULIO"]


# --- reporte_faltantes_completitud_xlsx: .xlsx descargable con las mismas
# facturas que ya se mostraron en la tabla de faltantes (pedido explícito
# del usuario, agosto 2026: número de factura, fecha, NIT y nombre del
# emisor, y valor -- en ese orden). ---

def test_reporte_faltantes_trae_numero_fecha_nit_nombre_y_valor():
    faltantes = [
        {"cufe": "CUFE-1", "fecha": "2026-07-15", "folio": "2", "prefijo": "FE",
         "proveedor_nit": "900111222", "proveedor_nombre": "PROVEEDOR TEST", "total": 100000},
        {"cufe": "CUFE-2", "fecha": "2026-07-20", "folio": "3", "prefijo": None,
         "proveedor_nit": "900333444", "proveedor_nombre": "OTRO PROVEEDOR", "total": 250000.5},
    ]

    contenido = orquestador.reporte_faltantes_completitud_xlsx(faltantes)

    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    assert filas[0] == ("Número de factura", "Fecha", "NIT emisor", "Nombre emisor", "Valor")
    assert filas[1][0] == "FE2"
    assert filas[1][1].date() == date(2026, 7, 15)  # openpyxl relee celdas de fecha como datetime
    assert filas[1][2] == "900111222"
    assert filas[1][3] == "PROVEEDOR TEST"
    assert filas[1][4] == 100000
    assert filas[2][0] == "3"  # sin prefijo -- no se inventa uno
    assert filas[2][2] == "900333444"
    assert filas[2][3] == "OTRO PROVEEDOR"
    assert filas[2][4] == 250000.5


def test_reporte_faltantes_vacio_da_solo_encabezado():
    contenido = orquestador.reporte_faltantes_completitud_xlsx([])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    filas = list(wb.active.iter_rows(values_only=True))
    assert filas == [("Número de factura", "Fecha", "NIT emisor", "Nombre emisor", "Valor")]


def test_encabezado_fecha_con_caracter_corrupto_igual_se_reconoce(empresa_configurada):
    """Confirmado contra un archivo real: el listado real trae 'Fecha
    Emisión' con el carácter de reemplazo U+FFFD en vez de 'ó'."""
    slug, tmp_path = empresa_configurada
    ruta_carpeta = tmp_path / "data" / "entrada-dian" / slug / "2026" / "07"
    ruta_carpeta.mkdir(parents=True)
    _crear_listado(
        ruta_carpeta / "listado.xlsx", [_fila("CUFE-1", "1", "900000000")], encabezado_fecha_corrupto=True,
    )

    r = orquestador.validar_completitud(slug, "2026/07", "listado.xlsx", "2026-07-01", "2026-07-31")

    assert r["total_listado"] == 1


def test_sin_facturas_del_periodo_en_el_listado_da_error_claro(empresa_configurada):
    slug, tmp_path = empresa_configurada
    ruta_carpeta = tmp_path / "data" / "entrada-dian" / slug / "2026" / "07"
    ruta_carpeta.mkdir(parents=True)
    _crear_listado(ruta_carpeta / "listado.xlsx", [_fila("CUFE-1", "1", "800000000")])  # NIT distinto

    with pytest.raises(ValueError, match="no trae ninguna factura recibida"):
        orquestador.validar_completitud(slug, "2026/07", "listado.xlsx", "2026-07-01", "2026-07-31")


def test_listar_archivos_listado_solo_lista_xlsx(empresa_configurada):
    slug, tmp_path = empresa_configurada
    ruta_carpeta = tmp_path / "data" / "entrada-dian" / slug / "2026" / "07"
    ruta_carpeta.mkdir(parents=True)
    _crear_listado(ruta_carpeta / "listado.xlsx", [_fila("CUFE-1", "1", "900000000")])
    (ruta_carpeta / "factura.zip").write_bytes(b"no es un xlsx")

    archivos = orquestador.listar_archivos_listado(slug, "2026/07")

    assert archivos == ["listado.xlsx"]


def test_listar_archivos_listado_carpeta_inexistente_devuelve_vacio(empresa_configurada):
    slug, _ = empresa_configurada
    assert orquestador.listar_archivos_listado(slug, "2026/07") == []


# --- validar_completitud_archivo_subido: SaaS, sin acceso al disco del servidor ---

def _bytes_listado(tmp_path, filas, nombre="listado-subido.xlsx"):
    """openpyxl solo escribe a disco -- se arma en un archivo temporal
    aparte (fuera de data/entrada-dian/) y se leen sus bytes, simulando
    justo lo que llega en request.files desde el navegador."""
    ruta_temp = tmp_path / nombre
    _crear_listado(ruta_temp, filas)
    contenido = ruta_temp.read_bytes()
    ruta_temp.unlink()
    return contenido


def test_validar_completitud_archivo_subido_funciona_igual_que_el_normal(empresa_configurada):
    slug, tmp_path = empresa_configurada
    filas = [_fila("CUFE-YA-IMPORTADA", "1", "900000000"), _fila("CUFE-FALTANTE", "2", "900000000")]
    contenido = _bytes_listado(tmp_path, filas)
    _sembrar_factura("CUFE-YA-IMPORTADA", "2026-07-01")

    r = orquestador.validar_completitud_archivo_subido(slug, "listado.xlsx", contenido, "2026-07-01", "2026-07-31")

    assert r["total_listado"] == 2
    assert [f["cufe"] for f in r["faltantes"]] == ["CUFE-FALTANTE"]


def test_validar_completitud_archivo_subido_lo_deja_guardado_para_despues(empresa_configurada):
    slug, tmp_path = empresa_configurada
    contenido = _bytes_listado(tmp_path, [_fila("CUFE-1", "1", "900000000")])

    orquestador.validar_completitud_archivo_subido(slug, "mi-listado.xlsx", contenido, "2026-07-01", "2026-07-31")

    carpeta = tmp_path / "data" / "entrada-dian" / slug / "_listados_subidos"
    guardados = list(carpeta.glob("*mi-listado.xlsx"))
    assert len(guardados) == 1


def test_validar_completitud_archivo_subido_sin_rango_da_error_claro(empresa_configurada):
    slug, tmp_path = empresa_configurada
    contenido = _bytes_listado(tmp_path, [_fila("CUFE-1", "1", "900000000")])

    with pytest.raises(ValueError, match="rango de fechas"):
        orquestador.validar_completitud_archivo_subido(slug, "listado.xlsx", contenido, "", "")


def test_validar_completitud_archivo_subido_sanea_nombre_con_ruta(empresa_configurada):
    slug, tmp_path = empresa_configurada
    contenido = _bytes_listado(tmp_path, [_fila("CUFE-1", "1", "900000000")])

    orquestador.validar_completitud_archivo_subido(
        slug, "../../../etc/malicioso.xlsx", contenido, "2026-07-01", "2026-07-31",
    )

    carpeta = tmp_path / "data" / "entrada-dian" / slug / "_listados_subidos"
    assert not (tmp_path / "etc").exists()
    assert list(carpeta.glob("*malicioso.xlsx"))
