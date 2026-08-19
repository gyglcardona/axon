"""
Lógica de orquestación, separada de cómo se invoca (CLI hoy, API/web mañana --
ver docs/00-contexto/decisiones-arquitectura.md, "Backend separado del
frontend"). `main.py` (CLI) y `backend/app.py` (API web) llaman a las mismas
funciones de aquí -- ninguno de los dos duplica la lógica de negocio.
"""

from __future__ import annotations

import datetime
import io
import json
import re
import sqlite3
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import auth
import auth_store
import contai_export
import drive_client
import gmail_client
import google_conexiones
import motor_sugerencias
import siigo_client
import siigo_payload
import state_store
from dian_parser import extraer_cufe, extraer_tercero, parsear_factura
from motor_reglas import clasificar_factura
from zip_handler import descubrir_documentos, DocumentoConError

TIPOS_CATALOGO_SIIGO = ("document_types", "payment_types", "journals", "taxes")
CAMPOS_ITEM_REPLICABLES = ("cuenta_contable", "iva_tax_id", "retencion_tax_id")
# Ítems que motor_reglas inyecta (no vienen de una línea real del XML) --
# "otros_impuestos" (impuestos que el parser no supo nombrar) y
# "politica_empresa" (IVA no discriminado de Hielo Super-Cool). Ninguno lleva
# IVA/retención propios, y ambos heredan la cuenta contable de las líneas de
# gasto reales (origen="xml") del mismo documento -- ver _aplicar_sugerencias.
_ORIGENES_INYECTADOS = {"otros_impuestos", "politica_empresa"}

REGISTRO = Path("config/empresas/registro.json")
CONFIG_EMPRESAS_DIR = Path("config/empresas")
CONFIG_PROVEEDORES_DIR = Path("config/proveedores")
ENTRADA_DIAN = Path("data/entrada-dian")
BASE_DATOS_EMPRESAS = Path("data/empresas")


class EmpresaNoEncontrada(Exception):
    pass


def _leer_registro() -> list[dict]:
    with open(REGISTRO, encoding="utf-8") as f:
        return json.load(f)["empresas"]


def resolver_empresa(slug: str) -> dict:
    for empresa in _leer_registro():
        if empresa["slug"] == slug:
            return empresa
    disponibles = ", ".join(e["slug"] for e in _leer_registro())
    raise EmpresaNoEncontrada(
        f"No existe la empresa '{slug}'. Slugs disponibles: {disponibles}. "
        "Nunca se adivina el NIT -- agrega la empresa al registro primero."
    )


def _leer_registro_json() -> dict:
    with open(REGISTRO, encoding="utf-8") as f:
        return json.load(f)


def _escribir_registro_json(datos: dict) -> None:
    with open(REGISTRO, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)
        f.write("\n")


_NIT_RE = re.compile(r"[^0-9]")
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

LIMITE_INTENTOS_REGISTRO_EMPRESA = 3
VENTANA_INTENTOS_REGISTRO_EMPRESA_HORAS = 1


def _validar_nit(nit_crudo: str) -> str:
    """Normaliza (solo dígitos, sin puntos/DV/espacios) y valida el NIT que
    llega de un formulario público sin sesión -- es la única defensa contra
    path traversal, porque el NIT termina siendo literalmente el nombre de
    archivo de `config/empresas/<nit>.json` y `data/empresas/<nit>.db`.
    Nunca se usa `nit_crudo` directamente en una ruta."""
    solo_digitos = _NIT_RE.sub("", nit_crudo or "")
    if not (5 <= len(solo_digitos) <= 15):
        raise ValueError("El NIT debe tener solo números (sin puntos ni dígito de verificación).")
    return solo_digitos


def _validar_email_publico(email: str) -> str:
    email = (email or "").strip().lower()
    if not _EMAIL_RE.match(email):
        raise ValueError("El correo no tiene un formato válido.")
    return email


def _generar_slug(nombre: str, slugs_existentes: set[str]) -> str:
    descompuesto = unicodedata.normalize("NFKD", nombre)
    solo_ascii = descompuesto.encode("ascii", "ignore").decode("ascii").lower()
    base = re.sub(r"[^a-z0-9]+", "-", solo_ascii).strip("-")[:60] or "empresa"
    slug, sufijo = base, 2
    while slug in slugs_existentes:
        slug = f"{base}-{sufijo}"
        sufijo += 1
    return slug


def _crear_md_empresa(nit: str, razon_social: str, slug: str) -> None:
    """Genera config/empresas/<nit>.md con lo ya conocido -- CLAUDE.md exige
    que toda empresa real tenga uno (regla 4). El resto (contacto contable,
    políticas activas) queda pendiente de completar a mano; nunca se
    sobreescribe si ya existe."""
    ruta_md = CONFIG_EMPRESAS_DIR / f"{nit}.md"
    if ruta_md.exists():
        return
    contenido = (
        f"# Conexión — {razon_social}\n\n"
        "> Datos generales y de conexión de esta empresa. **Nunca escribas aquí "
        "`usuario` ni `access_key` de Siigo** -- esos viven únicamente en "
        f"`config/empresas/{nit}.json`, que está en `.gitignore`.\n\n"
        "## Datos generales\n\n"
        f"- **Razón social:** {razon_social}\n"
        f"- **NIT:** {nit}\n"
        f"- **Slug:** {slug}\n"
        "- **Contacto contable:** (autorregistrada desde la web -- pendiente de completar)\n"
        "- **Régimen / responsabilidades tributarias relevantes:** (pendiente)\n\n"
        "## Notas operativas\n\n"
        "Empresa creada por autorregistro público -- sin políticas contables propias "
        "todavía. El motor de reglas opera con perfil de proveedor / regla genérica "
        "hasta que alguien configure sus políticas particulares (ver "
        "docs/02-reglas-negocio/README.md).\n"
    )
    CONFIG_EMPRESAS_DIR.mkdir(parents=True, exist_ok=True)
    ruta_md.write_text(contenido, encoding="utf-8")


def _crear_empresa_y_usuario_dueno(
    conn, nit: str, razon_social: str, email: str, creado_por_usuario_id: int | None = None,
) -> tuple[str, int]:
    """Mecánica compartida entre el autorregistro público
    (`registrar_empresa_nueva`) y la creación administrada
    (`crear_empresa_administrada`): valida que el NIT no exista todavía,
    crea el usuario dueño (rol "empresa", sin contraseña hasta que
    confirme), el registro de la empresa y materializa su base de datos
    aislada. No envía la invitación por correo -- cada llamador decide el
    texto exacto según quién la está creando.

    `creado_por_usuario_id`: quién queda como "creador" de ese usuario
    dueño para efectos de `auth._validar_puede_gestionar` -- en
    autorregistro nadie más lo creó (queda `None`); en creación
    administrada es quien la dio de alta, para que después pueda
    reenviarle la invitación o desactivarlo desde 'Usuarios' sin ser
    superusuario.

    No crea `config/empresas/<nit>.json` (credenciales/políticas): ese
    archivo se sigue creando solo, perezosamente, la primera vez que alguien
    guarda algo desde "Configuración" -- el motor de reglas ya sabe operar
    sin políticas propias (cae a perfil de proveedor / regla genérica)."""
    datos_registro = _leer_registro_json()
    if any(e["nit"] == nit for e in datos_registro["empresas"]):
        raise ValueError(
            f"Ya existe una empresa registrada con el NIT {nit}. Si es la tuya, pide a quien la "
            "administra que te invite desde 'Gestión de usuarios' en vez de crearla de nuevo."
        )

    usuario_id = auth_store.crear_usuario(
        conn, email=email, rol="empresa", puede_crear_usuarios=False,
        creado_por_usuario_id=creado_por_usuario_id,
    )
    auth_store.asociar_empresa_a_usuario(conn, usuario_id, nit)

    slugs_existentes = {e["slug"] for e in datos_registro["empresas"]}
    slug = _generar_slug(razon_social, slugs_existentes)
    datos_registro["empresas"].append({
        "slug": slug, "nit": nit, "nombre": razon_social, "config": f"config/empresas/{nit}.json",
    })
    _escribir_registro_json(datos_registro)
    _crear_md_empresa(nit, razon_social, slug)
    state_store.conectar(nit).close()  # materializa data/empresas/<nit>.db desde ya
    return slug, usuario_id


def _validar_datos_empresa_nueva(nit: str, razon_social: str, email: str) -> tuple[str, str, str]:
    nit = _validar_nit(nit)
    email = _validar_email_publico(email)
    razon_social = (razon_social or "").strip()
    if not razon_social:
        raise ValueError("La razón social es obligatoria.")
    if len(razon_social) > 200:
        raise ValueError("La razón social es demasiado larga (máximo 200 caracteres).")
    return nit, razon_social, email


def registrar_empresa_nueva(nit: str, razon_social: str, email: str) -> dict:
    """Autorregistro público de una empresa nueva -- la puerta de entrada al
    SaaS sin que un superusuario o contador tenga que invitarla primero.
    El rol del usuario creado siempre es "empresa" y nunca viene del cliente
    -- autorregistrarse jamás debe poder crear un superusuario o contador.

    Reutiliza el mismo mecanismo de token de una invitación normal
    (`auth.enviar_invitacion`): la cuenta queda sin contraseña hasta que se
    confirma el correo, así que nunca se puede iniciar sesión con una
    empresa que nadie verificó."""
    nit, razon_social, email = _validar_datos_empresa_nueva(nit, razon_social, email)

    conn = auth_store.conectar()
    try:
        intentos = auth_store.contar_intentos_registro_empresa_recientes(
            conn, email, VENTANA_INTENTOS_REGISTRO_EMPRESA_HORAS,
        )
        if intentos >= LIMITE_INTENTOS_REGISTRO_EMPRESA:
            raise ValueError(
                "Demasiados intentos de registro con este correo -- espera una hora e inténtalo de nuevo."
            )
        auth_store.registrar_intento_registro_empresa(conn, email)

        slug, usuario_id = _crear_empresa_y_usuario_dueno(conn, nit, razon_social, email)

        auth.enviar_invitacion(
            conn, usuario_id, email,
            "Confirma tu correo y crea tu contraseña -- AXON",
            f"Registraste {razon_social} en AXON.",
        )
    finally:
        conn.close()

    return {"registrado": True, "slug": slug, "nit": nit}


def crear_empresa_administrada(nit: str, razon_social: str, email: str, actor: dict) -> dict:
    """Crea una empresa nueva desde dentro del sistema -- para cuando un
    superusuario o un contador (con `puede_crear_usuarios`) necesita darla
    de alta él mismo, sin que el cliente tenga que autorregistrarse. A
    diferencia de `registrar_empresa_nueva`, quien crea NO es el dueño de
    la empresa: nunca ve ni necesita confirmar el correo del cliente -- el
    enlace para fijar la contraseña le llega directo a esa bandeja y el
    cliente lo usa cuando quiera, sin bloquear a quien la creó.

    Si `actor` es un contador (no superusuario), la empresa queda asociada
    también a él (además del usuario dueño que se crea) -- así puede
    gestionarla de una vez, y ningún otro contador o usuario sin
    asociación explícita la ve. Un superusuario nunca necesita esa
    asociación: ya ve todas las empresas por diseño. Sin el límite de
    intentos por hora del autorregistro público -- quien llama ya está
    autenticado y con permiso verificado, no es un formulario anónimo."""
    if not auth.puede_crear_empresas(actor):
        raise auth.AuthError("No tienes permiso para crear empresas nuevas.")
    nit, razon_social, email = _validar_datos_empresa_nueva(nit, razon_social, email)

    conn = auth_store.conectar()
    try:
        slug, usuario_id = _crear_empresa_y_usuario_dueno(
            conn, nit, razon_social, email, creado_por_usuario_id=actor["id"],
        )
        if actor["rol"] != "superusuario":
            auth_store.asociar_empresa_a_usuario(conn, actor["id"], nit)

        auth.enviar_invitacion(
            conn, usuario_id, email,
            "Te registraron en AXON -- confirma tu correo y crea tu contraseña",
            f"{actor['email']} registró {razon_social} en AXON.",
        )
    finally:
        conn.close()

    return {"creado": True, "slug": slug, "nit": nit}


def eliminar_empresa(slug: str) -> dict:
    """Elimina una empresa por completo -- su entrada en registro.json, su
    config/empresas/<nit>.json (credenciales/políticas), su .md de
    referencia, su base de datos aislada (TODA su causación importada) y
    cualquier acceso que algún usuario tuviera a ese NIT. No hay deshacer;
    la protección real vive en el endpoint (@requiere_superusuario en
    backend/app.py), esta función no vuelve a preguntar.

    Deliberadamente NO borra `data/entrada-dian/<slug>/` -- son los ZIP
    originales que entregó la empresa, con valor de respaldo/auditoría
    propio más allá de lo que AXON procesó con ellos."""
    empresa = resolver_empresa(slug)
    nit = empresa["nit"]

    datos_registro = _leer_registro_json()
    datos_registro["empresas"] = [e for e in datos_registro["empresas"] if e["nit"] != nit]
    _escribir_registro_json(datos_registro)

    for ruta in (
        CONFIG_EMPRESAS_DIR / f"{nit}.json",
        CONFIG_EMPRESAS_DIR / f"{nit}.md",
        BASE_DATOS_EMPRESAS / f"{nit}.db",
    ):
        ruta.unlink(missing_ok=True)

    conn = auth_store.conectar()
    try:
        auth_store.quitar_nit_de_todos_los_usuarios(conn, nit)
    finally:
        conn.close()

    return {"eliminada": True, "nit": nit, "slug": slug}


def listar_empresas() -> list[dict]:
    """Empresas del registro + cuántas facturas tiene cada una ya importadas
    + su destino de causación (para que la barra lateral sepa qué menús
    mostrar sin una llamada aparte por cada cambio de empresa activa)."""
    resultado = []
    for empresa in _leer_registro():
        db_path = BASE_DATOS_EMPRESAS / f"{empresa['nit']}.db"
        total = 0
        if db_path.exists():
            conn = sqlite3.connect(db_path)
            total = conn.execute("SELECT COUNT(*) FROM compras").fetchone()[0]
            conn.close()
        destino = obtener_destino_causacion(empresa["slug"])["destino_causacion"]
        resultado.append({**empresa, "total_facturas": total, "destino_causacion": destino})
    return resultado


def listar_carpetas_disponibles(slug: str) -> list[str]:
    """Subcarpetas <yyyy>/<mm> ya existentes bajo data/entrada-dian/<slug>/,
    para que la interfaz ofrezca un selector en vez de pedir una ruta escrita
    a mano (ver docs/03-ingesta-dian/carpetas-entrada.md)."""
    resolver_empresa(slug)  # valida que exista, sin usar el resultado
    base = ENTRADA_DIAN / slug
    if not base.exists():
        return []
    carpetas = []
    for año_dir in sorted(p for p in base.iterdir() if p.is_dir()):
        for mes_dir in sorted(p for p in año_dir.iterdir() if p.is_dir()):
            carpetas.append(f"{año_dir.name}/{mes_dir.name}")
    return carpetas


def listar_archivos_listado(slug: str, carpeta_relativa: str) -> list[str]:
    """Nombres de archivo `.xlsx` encontrados directo en
    data/entrada-dian/<slug>/<carpeta_relativa>/ (no recursivo) -- ahí es
    donde ya se espera que quede el listado de compras que exporta el
    portal de la DIAN para ese periodo, junto al ZIP (ver
    docs/03-ingesta-dian/carpetas-entrada.md)."""
    resolver_empresa(slug)
    carpeta = ENTRADA_DIAN / slug / carpeta_relativa
    if not carpeta.is_dir():
        return []
    return sorted(p.name for p in carpeta.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx")


def listar_facturas(slug: str) -> list[dict]:
    """Facturas ya importadas de una empresa, con ítems e impuestos, leídas en
    vivo de data/empresas/<nit>.db -- ver docs/05-esquema-datos/modelo-datos.md."""
    empresa = resolver_empresa(slug)
    db_path = BASE_DATOS_EMPRESAS / f"{empresa['nit']}.db"
    if not db_path.exists():
        return []

    # state_store.conectar (no sqlite3.connect directo) para garantizar que
    # corre la migración de columnas nuevas (ver state_store._migrar) aunque
    # esta sea la primera llamada de la sesión contra esta base.
    conn = state_store.conectar(empresa["nit"])
    conn.row_factory = sqlite3.Row
    try:
        compras = conn.execute(
            "SELECT * FROM compras ORDER BY fecha_emision DESC, id DESC"
        ).fetchall()
        proveedores_autorretenedores: dict[str, bool] = {}
        salida = []
        for c in compras:
            items = conn.execute(
                "SELECT * FROM detalle_compras WHERE compra_id = ? ORDER BY orden", (c["id"],)
            ).fetchall()
            items_out = []
            for it in items:
                impuestos = conn.execute(
                    "SELECT tipo, porcentaje, valor FROM detalle_impuestos WHERE detalle_compra_id = ?",
                    (it["id"],),
                ).fetchall()
                items_out.append({
                    "id": it["id"], "descripcion": it["descripcion"], "cantidad": it["cantidad"],
                    "valor_unitario": it["valor_unitario"], "cuenta_contable": it["cuenta_contable"],
                    "iva_tax_id": it["iva_tax_id"], "retencion_tax_id": it["retencion_tax_id"],
                    "descuento_monto": it["descuento_monto"] or 0, "tipo_item": it["tipo_item"],
                    "origen": it["origen"], "impuestos": [dict(i) for i in impuestos],
                })
            nit = c["proveedor_nit"]
            if nit not in proveedores_autorretenedores:
                proveedores_autorretenedores[nit] = motor_sugerencias.es_autorretenedor(nit)
            salida.append({
                "cufe": c["cufe"], "numero_factura": c["numero_factura"],
                "prefijo": c["prefijo"], "numero_puro": c["numero_puro"], "fecha_emision": c["fecha_emision"],
                "proveedor_nit": nit, "proveedor_nombre": c["proveedor_nombre"],
                "proveedor_autorretenedor": proveedores_autorretenedores[nit],
                "subtotal_xml": c["subtotal_xml"], "total_pagar_xml": c["total_pagar_xml"],
                "resuelto_por": c["resuelto_por"], "estado_siigo": c["estado_siigo"],
                "estado_contai": c["estado_contai"],
                "tipo_comprobante_id": c["tipo_comprobante_id"], "medio_pago_id": c["medio_pago_id"],
                "modo_pago_contai": c["modo_pago_contai"],
                "siigo_id": c["siigo_id"], "siigo_error": c["siigo_error"],
                "archivo_origen": c["archivo_origen"],
                "notas": json.loads(c["notas"]) if c["notas"] else [],
                "items": items_out,
            })
        return salida
    finally:
        conn.close()


def _ruta_config_empresa(empresa: dict) -> Path:
    return CONFIG_EMPRESAS_DIR / f"{empresa['nit']}.json"


def obtener_conexion_siigo(slug: str) -> dict:
    """Datos de conexión a Siigo de config/empresas/<nit>.json -- hoy vienen
    precargados ahí para las empresas locales, pero este mismo formulario es
    con el que una empresa nueva del SaaS los llenaría por primera vez (no
    tendrá ningún archivo todavía)."""
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)
    if not ruta.exists():
        return {"usuario": "", "access_key": "", "partner_id": "", "configurado": False}

    with open(ruta, encoding="utf-8") as f:
        config = json.load(f)
    cred = config.get("credenciales_siigo", {})
    usuario = cred.get("usuario", "")
    access_key = cred.get("access_key", "")
    return {
        "usuario": usuario,
        "access_key": access_key,
        "partner_id": cred.get("partner_id", ""),
        "configurado": bool(usuario and access_key),
    }


def guardar_conexion_siigo(slug: str, usuario: str, access_key: str, partner_id: str) -> dict:
    """Guarda usuario/access_key/partner_id en config/empresas/<nit>.json.
    Si el archivo no existe todavía (empresa nueva sin config previa), lo crea
    -- pero nunca inventa políticas contables, esas se agregan aparte cuando
    se confirmen con la contadora (ver docs/07-operacion-claude-code/comandos.md)."""
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)

    if ruta.exists():
        with open(ruta, encoding="utf-8") as f:
            config = json.load(f)
    else:
        config = {"nit": empresa["nit"], "nombre": empresa["nombre"], "slug": slug,
                   "credenciales_siigo": {}, "politicas": {}}

    config.setdefault("credenciales_siigo", {})
    config["credenciales_siigo"]["usuario"] = usuario
    config["credenciales_siigo"]["access_key"] = access_key
    config["credenciales_siigo"]["partner_id"] = partner_id

    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
        f.write("\n")

    return {"guardado": True}


def obtener_conexion_drive(slug: str) -> dict:
    """Carpeta de Google Drive configurada para esta empresa, de
    config/empresas/<nit>.json, y qué conexión de Google se usa para leerla
    (`conexion_id`, "" = la conexión legacy compartida -- ver
    google_conexiones.py)."""
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)
    if not ruta.exists():
        return {"carpeta_id": "", "configurado": False, "conexion_id": ""}

    with open(ruta, encoding="utf-8") as f:
        config = json.load(f)
    conexion_drive = config.get("conexion_drive", {})
    carpeta_id = conexion_drive.get("carpeta_id", "")
    return {
        "carpeta_id": carpeta_id, "configurado": bool(carpeta_id),
        "conexion_id": conexion_drive.get("conexion_id", ""),
    }


def guardar_conexion_drive(slug: str, carpeta_id: str) -> dict:
    """Guarda el id de la carpeta de Drive de esta empresa en
    config/empresas/<nit>.json -- mismo patrón que guardar_conexion_siigo."""
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)

    if ruta.exists():
        with open(ruta, encoding="utf-8") as f:
            config = json.load(f)
    else:
        config = {"nit": empresa["nit"], "nombre": empresa["nombre"], "slug": slug,
                   "credenciales_siigo": {}, "conexion_drive": {}, "politicas": {}}

    config.setdefault("conexion_drive", {})
    config["conexion_drive"]["carpeta_id"] = carpeta_id

    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
        f.write("\n")

    return {"guardado": True}


def listar_conexiones_google() -> list[dict]:
    return google_conexiones.listar_conexiones()


def asociar_conexion_google(slug: str, conexion_id: str) -> dict:
    """Asocia una conexión de Google ya existente (creada por OAuth o la
    legacy) a esta empresa -- reutilizar sin repetir el consentimiento."""
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)

    if ruta.exists():
        with open(ruta, encoding="utf-8") as f:
            config = json.load(f)
    else:
        config = {"nit": empresa["nit"], "nombre": empresa["nombre"], "slug": slug,
                   "credenciales_siigo": {}, "conexion_drive": {}, "politicas": {}}

    config.setdefault("conexion_drive", {})
    config["conexion_drive"]["conexion_id"] = conexion_id

    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
        f.write("\n")

    return {"guardado": True}


def _nits_asociados_a_conexion(conexion_id: str) -> set[str]:
    """Qué NITs ya usan esta conexión de Google (la tienen guardada en su
    propio `conexion_drive.conexion_id`) -- sirve para decidir si un usuario
    sin `puede_crear_usuarios`/superusuario puede ver una conexión que no
    creó él mismo pero que ya está en uso en una empresa suya."""
    nits = set()
    for empresa in _leer_registro():
        if obtener_conexion_drive(empresa["slug"])["conexion_id"] == conexion_id:
            nits.add(empresa["nit"])
    return nits


def listar_conexiones_google_visibles(usuario: dict) -> list[dict]:
    """Filtra `listar_conexiones_google()` según quién pregunta -- cierra el
    gap que motivó la Sesión 8 del plan de login (antes cualquier usuario
    logueado veía las conexiones de TODAS las empresas, no solo las suyas).
    Superusuario ve todas; cualquier otro rol solo ve las que él mismo creó
    o las que ya están asociadas a una empresa que puede ver."""
    todas = listar_conexiones_google()
    if usuario["rol"] == "superusuario":
        return todas

    conn = auth_store.conectar()
    try:
        nits_visibles = set(auth_store.listar_nits_de_usuario(conn, usuario["id"]))
    finally:
        conn.close()

    visibles = []
    for conexion in todas:
        if conexion.get("creado_por_usuario_id") == usuario["id"]:
            visibles.append(conexion)
        elif _nits_asociados_a_conexion(conexion["id"]) & nits_visibles:
            visibles.append(conexion)
    return visibles


def usuario_puede_usar_conexion_google(usuario: dict, conexion_id: str) -> bool:
    """¿Puede `usuario` asociar esta conexión a una empresa suya? Misma
    regla que `listar_conexiones_google_visibles` -- evita que alguien
    asocie a su empresa una conexión ajena solo por conocer/adivinar su id."""
    if not conexion_id:
        return True  # "" = usar la conexión legacy por defecto, comportamiento de siempre
    ids_visibles = {c["id"] for c in listar_conexiones_google_visibles(usuario)}
    return conexion_id in ids_visibles


def asociar_conexion_google_para_usuario(slug: str, conexion_id: str, usuario: dict) -> dict:
    """Como `asociar_conexion_google`, pero antes valida que `usuario` tenga
    visibilidad sobre `conexion_id` (ver `usuario_puede_usar_conexion_google`)."""
    if not usuario_puede_usar_conexion_google(usuario, conexion_id):
        raise ValueError("No tienes acceso a esa conexión de Google.")
    return asociar_conexion_google(slug, conexion_id)

    return {"guardado": True}


def iniciar_autorizacion_google(slug: str, usuario_id: int | None = None) -> str:
    """Devuelve la URL de consentimiento de Google para conectar una cuenta
    nueva a nombre de `slug` -- ver google_conexiones.iniciar_autorizacion.
    `usuario_id` queda como dueño de la conexión nueva (ver
    listar_conexiones_google_visibles)."""
    resolver_empresa(slug)  # valida que la empresa exista antes de iniciar el flujo
    return google_conexiones.iniciar_autorizacion(slug, usuario_id)


def completar_autorizacion_google(state: str, code: str) -> dict:
    """Completa el flujo OAuth y asocia la conexión nueva automáticamente a
    la empresa que lo inició (conectar para una empresa la deja lista de
    una vez, sin paso manual extra)."""
    resultado = google_conexiones.procesar_callback(state, code)
    asociar_conexion_google(resultado["slug"], resultado["conexion_id"])
    return resultado


_CONFIG_GMAIL_DEFAULT = {
    "activo": False,
    "buscar_en_spam": True,
    "desde_fecha": "",
    "ultima_sincronizacion": "",
}


def obtener_conexion_gmail(slug: str) -> dict:
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)
    if not ruta.exists():
        return dict(_CONFIG_GMAIL_DEFAULT)

    with open(ruta, encoding="utf-8") as f:
        config = json.load(f)
    guardado = config.get("conexion_gmail") or {}
    return {**_CONFIG_GMAIL_DEFAULT, **guardado}


def guardar_conexion_gmail(slug: str, campos: dict) -> dict:
    """Guarda la configuración de ingesta desde Gmail. `desde_fecha` se fija
    una sola vez -- si la empresa ya tiene una guardada, se ignora un valor
    nuevo (regla de negocio confirmada: "se fija una vez, luego solo
    avanza"). Bloquea activar Gmail mientras la empresa siga en la conexión
    legacy: esa conexión nunca se autorizó con el scope de Gmail (se creó
    antes de que existiera), activarla daría un error crudo de permisos de
    Google en vez de un mensaje claro."""
    conexion_drive = obtener_conexion_drive(slug)
    actual = obtener_conexion_gmail(slug)

    if campos.get("activo") and not conexion_drive["conexion_id"]:
        raise ValueError(
            "La conexión compartida actual no tiene permiso de Gmail (se autorizó antes de que "
            "existiera esta función) -- conecta una cuenta de Google nueva para esta empresa "
            "desde 'Configuración' antes de activar la importación desde Gmail."
        )

    nuevo = {**actual, **campos}
    if actual["desde_fecha"]:
        nuevo["desde_fecha"] = actual["desde_fecha"]  # ya fijada -- nunca se sobreescribe

    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)
    if ruta.exists():
        with open(ruta, encoding="utf-8") as f:
            config = json.load(f)
    else:
        config = {"nit": empresa["nit"], "nombre": empresa["nombre"], "slug": slug,
                   "credenciales_siigo": {}, "conexion_drive": {}, "politicas": {}}

    config["conexion_gmail"] = {**_CONFIG_GMAIL_DEFAULT, **config.get("conexion_gmail", {}), **nuevo}

    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
        f.write("\n")

    return {"guardado": True}


DESTINOS_CAUSACION_VALIDOS = ("siigo", "contai")


def obtener_destino_causacion(slug: str) -> dict:
    """A qué sistema se causa esta empresa -- "siigo" (API real) o "contai"
    (archivos planos, todavía sin implementar). Si el archivo de config no
    existe o no trae el campo, el default es "siigo" -- así ninguna empresa
    ya configurada cambia de comportamiento sin tocar su archivo."""
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)
    if not ruta.exists():
        return {"destino_causacion": "siigo"}

    with open(ruta, encoding="utf-8") as f:
        config = json.load(f)
    return {"destino_causacion": config.get("destino_causacion") or "siigo"}


def guardar_destino_causacion(slug: str, destino: str) -> dict:
    """Guarda el destino de causación de esta empresa en
    config/empresas/<nit>.json -- mismo patrón que guardar_conexion_drive."""
    if destino not in DESTINOS_CAUSACION_VALIDOS:
        raise ValueError(
            f"Destino de causación inválido: '{destino}'. Válidos: {', '.join(DESTINOS_CAUSACION_VALIDOS)}."
        )
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)

    if ruta.exists():
        with open(ruta, encoding="utf-8") as f:
            config = json.load(f)
    else:
        config = {"nit": empresa["nit"], "nombre": empresa["nombre"], "slug": slug,
                   "credenciales_siigo": {}, "conexion_drive": {}, "politicas": {}}

    config["destino_causacion"] = destino

    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
        f.write("\n")

    return {"guardado": True}


_CONFIG_CONTAI_DEFAULT = {
    "comprobante": "00010",
    "modo_pago_default": "contado",
    "cuenta_credito_contado": "",
    "cuenta_credito_credito": "",
    "cuentas_iva_por_tarifa": {},
    "cuentas_retencion_por_tipo": {},
    # Cuenta de GASTO (la BASE) por tarifa de IVA -- reemplaza la cuenta por
    # ítem/producto, agrupando el asiento por totales (pedido del contador,
    # agosto 2026: "los asientos se hacen en base a los totales, no en base
    # a cada ítem"). La clave "0.0" cubre las líneas sin IVA (ej. compras
    # excluidas/no gravadas) -- es una tarifa más, no una etiqueta especial.
    # Ver contai_export._categoria_gasto_de_linea.
    "cuentas_gasto_por_categoria": {},
    # Cuenta de IMPUESTO por tipo, para impuestos que no son IVA ni
    # retención (ej. "INC" = Impuesto al Consumo, "ICUI" = IBUA/ICUI, ver
    # dian_parser.TAX_SCHEME_MAP) -- el valor de ese impuesto va en su
    # propia fila de débito, aparte de la base (confirmado por el usuario,
    # agosto 2026: "va aparte, fila propia"). Ver
    # contai_export.construir_movimientos.
    "cuentas_impuesto_por_tipo": {},
}


def obtener_config_contai(slug: str) -> dict:
    """Catálogo de cuentas de Contai de esta empresa (comprobante, cuentas
    de IVA por tarifa, de retención por tipo, y de crédito según modo de
    pago) -- solo aplica si `destino_causacion` es "contai". Vacío/sin
    archivo → valores por defecto, nunca `None` (así `contai_export` siempre
    recibe la forma completa del dict, aunque esté sin configurar)."""
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)
    if not ruta.exists():
        return dict(_CONFIG_CONTAI_DEFAULT)

    with open(ruta, encoding="utf-8") as f:
        config = json.load(f)
    guardado = config.get("config_contai") or {}
    return {**_CONFIG_CONTAI_DEFAULT, **guardado}


def guardar_config_contai(slug: str, config_contai: dict) -> dict:
    """Guarda el catálogo de cuentas de Contai -- mismo patrón que
    guardar_destino_causacion. Nunca valida que las cuentas existan en el
    plan de cuentas importado (el plan de cuentas es de solo lectura /
    referencia, no una restricción dura -- igual que Siigo)."""
    empresa = resolver_empresa(slug)
    ruta = _ruta_config_empresa(empresa)

    if ruta.exists():
        with open(ruta, encoding="utf-8") as f:
            config = json.load(f)
    else:
        config = {"nit": empresa["nit"], "nombre": empresa["nombre"], "slug": slug,
                   "credenciales_siigo": {}, "conexion_drive": {}, "politicas": {}}

    config["config_contai"] = {**_CONFIG_CONTAI_DEFAULT, **config.get("config_contai", {}), **config_contai}

    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
        f.write("\n")

    return {"guardado": True}


def importar_desde_drive(slug: str) -> dict:
    """Sincroniza TODO lo nuevo de la carpeta de Drive de esta empresa hacia
    data/entrada-dian/<slug>/ (reflejando la misma estructura de subcarpetas
    que tenga en Drive -- no se asume año/mes, cada empresa la puede
    organizar como quiera, ver docs/03-ingesta-dian/importar-desde-drive.md)
    y luego corre el pipeline de importación normal sobre todo el árbol de
    la empresa. Nunca borra ni sobreescribe un archivo local ya descargado
    -- la deduplicación real la hace `ejecutar_importar` por CUFE, esto solo
    evita descargar de nuevo bytes que ya están en disco."""
    empresa = resolver_empresa(slug)
    conexion = obtener_conexion_drive(slug)
    if not conexion["configurado"]:
        raise ValueError(
            "Esta empresa todavía no tiene una carpeta de Drive configurada -- "
            "complétala primero en el menú 'Configuración'."
        )

    creds = google_conexiones.obtener_credenciales(conexion["conexion_id"])
    arbol = drive_client.listar_arbol(conexion["carpeta_id"], creds)
    carpeta_empresa = ENTRADA_DIAN / slug

    descargados, ya_estaban = 0, 0
    for archivo in arbol:
        destino = carpeta_empresa / archivo["ruta_relativa"]
        if destino.is_file():
            ya_estaban += 1
            continue
        destino.parent.mkdir(parents=True, exist_ok=True)
        contenido = drive_client.descargar_archivo(archivo["id"], creds)
        destino.write_bytes(contenido)
        descargados += 1

    resumen_importar = ejecutar_importar(slug, ".")
    return {"descargados": descargados, "ya_estaban_localmente": ya_estaban, **resumen_importar}


def importar_desde_gmail(slug: str) -> dict:
    """Sincroniza TODO lo nuevo desde Gmail (bandeja de entrada + spam según
    configuración) hacia data/entrada-dian/<slug>/gmail/<yyyy>/<mm>/, y luego
    corre el pipeline de importación normal sobre esa carpeta. El nombre de
    archivo incluye el `message_id` de Gmail -- eso ya garantiza que una
    corrida repetida no descargue dos veces el mismo adjunto (mismo criterio
    de idempotencia por existencia de archivo que usa Drive); la
    deduplicación real de facturas la sigue haciendo `ejecutar_importar` por
    CUFE."""
    empresa = resolver_empresa(slug)
    conexion_gmail = obtener_conexion_gmail(slug)
    if not conexion_gmail["activo"]:
        raise ValueError(
            "Esta empresa todavía no tiene activada la importación desde Gmail -- "
            "actívala primero en el menú 'Configuración'."
        )
    if not conexion_gmail["desde_fecha"]:
        raise ValueError(
            "Falta definir desde qué fecha buscar en Gmail -- configúralo en el menú 'Configuración'."
        )

    conexion_drive = obtener_conexion_drive(slug)
    creds = google_conexiones.obtener_credenciales(conexion_drive["conexion_id"])
    desde_fecha = conexion_gmail["ultima_sincronizacion"] or conexion_gmail["desde_fecha"]
    adjuntos = gmail_client.buscar_adjuntos_zip(creds, desde_fecha, conexion_gmail["buscar_en_spam"])

    carpeta_gmail = ENTRADA_DIAN / slug / "gmail"
    descargados, ya_estaban = 0, 0
    fecha_mas_reciente_ms = 0
    for adjunto in adjuntos:
        fecha_ms = int(adjunto["fecha_interna"] or 0)
        fecha_mas_reciente_ms = max(fecha_mas_reciente_ms, fecha_ms)
        fecha_msg = datetime.datetime.fromtimestamp(fecha_ms / 1000, tz=datetime.timezone.utc)
        subcarpeta = carpeta_gmail / f"{fecha_msg:%Y}" / f"{fecha_msg:%m}"
        destino = subcarpeta / f"{adjunto['message_id']}_{adjunto['filename']}"
        if destino.is_file():
            ya_estaban += 1
            continue
        subcarpeta.mkdir(parents=True, exist_ok=True)
        contenido = gmail_client.descargar_adjunto(creds, adjunto["message_id"], adjunto["attachment_id"])
        destino.write_bytes(contenido)
        descargados += 1

    if fecha_mas_reciente_ms:
        nueva_fecha = datetime.datetime.fromtimestamp(
            fecha_mas_reciente_ms / 1000, tz=datetime.timezone.utc
        ).strftime("%Y-%m-%d")
        guardar_conexion_gmail(slug, {"ultima_sincronizacion": nueva_fecha})

    resumen_importar = ejecutar_importar(slug, "gmail") if carpeta_gmail.is_dir() else {
        "empresa": empresa["nombre"], "nit": empresa["nit"], "carpeta": str(carpeta_gmail),
        "nuevas": 0, "ya_existentes": 0, "duplicados": 0, "no_facturas": 0, "con_error": 0,
        "nit_no_corresponde": 0,
    }
    return {"descargados": descargados, "ya_estaban_localmente": ya_estaban, **resumen_importar}


def importar_desde_google(slug: str) -> dict:
    """Sincroniza Drive (si tiene carpeta configurada) y Gmail (si está
    activo) en una sola corrida -- lo que llama el botón único del toolbar.
    Error claro si ninguno de los dos está configurado."""
    conexion_drive = obtener_conexion_drive(slug)
    conexion_gmail = obtener_conexion_gmail(slug)
    if not conexion_drive["configurado"] and not conexion_gmail["activo"]:
        raise ValueError(
            "Esta empresa no tiene ni Drive ni Gmail configurados -- configura al menos uno "
            "en el menú 'Configuración' antes de importar desde Google."
        )

    resultado = {"drive": None, "gmail": None}
    if conexion_drive["configurado"]:
        resultado["drive"] = importar_desde_drive(slug)
    if conexion_gmail["activo"]:
        resultado["gmail"] = importar_desde_gmail(slug)
    return resultado


def importar_todo(slug: str) -> dict:
    """El botón único de importación: trae lo nuevo de Drive/Gmail (si están
    configurados) y además corre el pipeline sobre TODO lo que ya esté en
    data/entrada-dian/<slug>/ -- así el mismo botón sirve por igual para una
    empresa que solo recibe el ZIP a mano (o por upload, ver
    guardar_archivos_subidos), una que usa Drive/Gmail, o ambas cosas a la
    vez, sin que quien lo usa tenga que saber cuál es su caso.

    Que Google no esté configurado, o que falle (token vencido, etc.), no
    debe tumbar la importación local -- por eso ese error se captura y se
    devuelve como dato (`error_google`) en vez de propagarse. La pasada
    local siempre corre al final y es idempotente (`ejecutar_importar`
    dedupea por CUFE), así que no duplica nada aunque Drive/Gmail ya hayan
    importado sus propios archivos nuevos momentos antes."""
    resultado: dict = {"local": None, "drive": None, "gmail": None, "error_google": None}

    conexion_drive = obtener_conexion_drive(slug)
    conexion_gmail = obtener_conexion_gmail(slug)
    if conexion_drive["configurado"] or conexion_gmail["activo"]:
        try:
            google = importar_desde_google(slug)
            resultado["drive"] = google["drive"]
            resultado["gmail"] = google["gmail"]
        except (ValueError, drive_client.DriveError, gmail_client.GmailError, google_conexiones.GoogleConexionError) as e:
            resultado["error_google"] = str(e)

    empresa = resolver_empresa(slug)
    carpeta_local = ENTRADA_DIAN / slug
    if carpeta_local.is_dir():
        resultado["local"] = ejecutar_importar(slug, ".")
    else:
        resultado["local"] = {
            "empresa": empresa["nombre"], "nit": empresa["nit"], "carpeta": str(carpeta_local),
            "nuevas": 0, "ya_existentes": 0, "duplicados": 0, "no_facturas": 0, "con_error": 0,
        }
    return resultado


def _ruta_relativa_segura(nombre_crudo: str) -> Path | None:
    """Convierte un nombre (o ruta relativa, ej. la que trae un <input
    webkitdirectory> con subcarpetas) en una ruta relativa segura para
    escribir dentro de una carpeta propia: nunca absoluta, nunca con '..',
    nunca con una letra de unidad de Windows ("C:"). Devuelve None si
    después de sanear no queda nada útil."""
    crudo = nombre_crudo.replace("\\", "/")
    partes = [p for p in Path(crudo).parts if p not in ("..", ".", "") and not p.endswith(":")]
    if not partes:
        return None
    return Path(*partes)


def guardar_archivos_subidos(slug: str, archivos: list[tuple[str, bytes]]) -> dict:
    """Guarda ZIP/XML subidos desde el navegador en
    data/entrada-dian/<slug>/subidos/<marca-de-tiempo>/ -- para el modo SaaS,
    donde quien importa no tiene ni sabe de una ruta en el servidor (ver
    docs/06-multiempresa-saas/). Quedan ahí de forma permanente, igual que
    cualquier otro archivo entregado por la empresa -- mismo criterio que el
    resto de data/entrada-dian/, nunca se borran solos.

    `nombre` puede traer subcarpetas (el selector de carpeta del navegador
    manda `webkitRelativePath`, ej. "2026/07/factura1.zip") -- se preservan,
    para no chocar nombres iguales que vivían en meses/carpetas distintas del
    árbol original. Se sanea con `_ruta_relativa_segura` para que ninguna
    ruta (ni una manipulada a mano contra la API) pueda escribir fuera de la
    carpeta de subidos."""
    resolver_empresa(slug)
    marca = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    carpeta = ENTRADA_DIAN / slug / "subidos" / marca
    carpeta.mkdir(parents=True, exist_ok=True)
    guardados = []
    for nombre, contenido in archivos:
        ruta_segura = _ruta_relativa_segura(nombre)
        if ruta_segura is None:
            continue
        destino = carpeta / ruta_segura
        destino.parent.mkdir(parents=True, exist_ok=True)
        destino.write_bytes(contenido)
        guardados.append(str(ruta_segura))
    return {"carpeta": str(carpeta), "archivos": guardados}


def actualizar_catalogos_siigo(slug: str) -> dict:
    """Autentica con las credenciales guardadas (menú Conexión Siigo) y trae
    los 4 catálogos maestros reales de Siigo: tipos de FC, medios de pago,
    comprobantes contables e impuestos/retenciones. Son de solo lectura --
    no crean, modifican ni envían nada en Siigo (ver CLAUDE.md, regla 3)."""
    conexion = obtener_conexion_siigo(slug)
    if not conexion["configurado"]:
        raise ValueError(
            "Esta empresa todavía no tiene usuario/access_key configurados -- "
            "complétalos primero en el menú 'Conexión Siigo'."
        )

    empresa = resolver_empresa(slug)
    token = siigo_client.autenticar(conexion["usuario"], conexion["access_key"])
    partner_id = conexion["partner_id"]

    catalogos = {
        "document_types": siigo_client.obtener_document_types(token, partner_id),
        "payment_types": siigo_client.obtener_payment_types(token, partner_id),
        "journals": siigo_client.obtener_journals(token, partner_id),
        "taxes": siigo_client.obtener_taxes(token, partner_id),
    }

    conn = state_store.conectar(empresa["nit"])
    try:
        return {tipo: state_store.guardar_catalogo_siigo(conn, tipo, items) for tipo, items in catalogos.items()}
    finally:
        conn.close()


def listar_catalogo_siigo(slug: str, tipo: str) -> list[dict]:
    if tipo not in TIPOS_CATALOGO_SIIGO:
        raise ValueError(f"Tipo de catálogo desconocido: '{tipo}'. Válidos: {', '.join(TIPOS_CATALOGO_SIIGO)}.")
    empresa = resolver_empresa(slug)
    db_path = BASE_DATOS_EMPRESAS / f"{empresa['nit']}.db"
    if not db_path.exists():
        return []
    conn = state_store.conectar(empresa["nit"])
    try:
        return state_store.listar_catalogo_siigo(conn, tipo)
    finally:
        conn.close()


def _mapear_compra_siigo(compra: dict, nombre_proveedor: str) -> dict:
    nit = compra.get("supplier", {}).get("identification", "")
    prov_inv = compra.get("provider_invoice") or {}
    factura_proveedor = f"{prov_inv.get('prefix') or ''}{prov_inv.get('number') or ''}"
    items = compra.get("items", [])
    subtotal = sum(it.get("total", 0) for it in items)
    pagos = compra.get("payments") or []
    return {
        "siigo_id": compra.get("id"),
        "numero": compra.get("number"),
        "fecha": compra.get("date"),
        "proveedor_nit": nit,
        "proveedor_nombre": nombre_proveedor,
        "factura_proveedor": factura_proveedor,
        "total": compra.get("total", 0),
        "subtotal": subtotal,
        # Tipo de comprobante y medio de pago REALES con los que ya se causó
        # esta compra en Siigo -- confirmado real: Siigo sí los devuelve en
        # GET /v1/purchases (`document.id`, `payments[0].id`), pero antes de
        # este cambio se descartaban al cachear. motor_sugerencias.
        # sugerir_cabecera los usa como histórico por proveedor (caso real:
        # Construcciones y Adecuaciones ET, causada antes por el aplicativo
        # anterior del usuario -- sin esto, ninguna factura nueva del mismo
        # proveedor podía autocompletar la cabecera).
        "tipo_comprobante_id": (compra.get("document") or {}).get("id"),
        "medio_pago_id": pagos[0].get("id") if pagos else None,
        "items": [
            {
                "descripcion": it.get("description"),
                "cuenta_contable": it.get("code"),
                "tipo": it.get("type"),
                "cantidad": it.get("quantity"),
                "valor_unitario": it.get("price"),
                "total": it.get("total"),
                "impuestos": [
                    {"tipo": tax.get("name"), "porcentaje": tax.get("percentage"), "valor": tax.get("value")}
                    for tax in it.get("taxes", [])
                ],
            }
            for it in items
        ],
    }


def descargar_compras_siigo(slug: str, desde: str | None = None, hasta: str | None = None) -> dict:
    """Trae las compras ya causadas en Siigo (GET /v1/purchases, paginado) y
    las guarda en el caché local -- botón 'Descargar'.

    Siigo no filtra por fecha en este endpoint (ver `siigo_client.obtener_purchases_pagina`),
    pero los resultados vienen ordenados por consecutivo/fecha descendente
    (el más reciente primero). Con `desde`/`hasta` se acota el trabajo real:
    se descartan las compras más recientes que `hasta` y, apenas aparece una
    compra más vieja que `desde`, se corta la paginación -- así se evita
    traer miles de páginas para consultar solo un rango reciente. Sin
    `desde` ni `hasta` se trae todo (mismo comportamiento de siempre) y se
    reemplaza el caché local completo; con rango, se hace upsert sobre el
    caché existente para no perder compras de otros rangos ya descargados."""
    conexion = obtener_conexion_siigo(slug)
    if not conexion["configurado"]:
        raise ValueError(
            "Esta empresa todavía no tiene usuario/access_key configurados -- "
            "complétalos primero en el menú 'Conexión Siigo'."
        )

    empresa = resolver_empresa(slug)
    token = siigo_client.autenticar(conexion["usuario"], conexion["access_key"])
    partner_id = conexion["partner_id"]

    MAX_PAGINAS = 200
    PAGE_SIZE = 100
    en_rango: list[dict] = []
    page = 1
    detenido_por_fecha = False
    while page <= MAX_PAGINAS:
        resultados, paginacion = siigo_client.obtener_purchases_pagina(token, partner_id, page, PAGE_SIZE)
        if not resultados:
            break
        for compra in resultados:
            fecha = compra.get("date") or ""
            if desde and fecha and fecha < desde:
                detenido_por_fecha = True
                break
            if hasta and fecha and fecha > hasta:
                continue  # más reciente que el rango pedido -- se descarta, se sigue paginando
            en_rango.append(compra)
        if detenido_por_fecha:
            break
        if len(resultados) < PAGE_SIZE:
            break  # página incompleta -- era la última, sin importar lo que diga total_results
        total = paginacion.get("total_results", 0)
        if page * PAGE_SIZE >= total:
            break
        page += 1

    conn = state_store.conectar(empresa["nit"])
    try:
        filas = []
        for compra in en_rango:
            nit = compra.get("supplier", {}).get("identification", "")
            nombre = state_store.obtener_nombre_proveedor_siigo(conn, nit) if nit else None
            if nombre is None and nit:
                nombre = siigo_client.obtener_nombre_proveedor(token, partner_id, nit) or nit
                state_store.guardar_nombre_proveedor_siigo(conn, nit, nombre)
            filas.append(_mapear_compra_siigo(compra, nombre or nit))

        reemplazar_todo = not desde and not hasta
        total_guardadas = state_store.guardar_compras_siigo(conn, filas, reemplazar_todo=reemplazar_todo)
    finally:
        conn.close()

    return {"total": total_guardadas}


def listar_compras_siigo(
    slug: str, desde: str | None = None, hasta: str | None = None, texto: str | None = None,
) -> list[dict]:
    """Lee el caché local de compras ya descargadas (ver `descargar_compras_siigo`)
    -- nunca llama a Siigo. Si todavía no se ha descargado nada, devuelve
    lista vacía (no es un error: es un estado inicial normal)."""
    empresa = resolver_empresa(slug)
    db_path = BASE_DATOS_EMPRESAS / f"{empresa['nit']}.db"
    if not db_path.exists():
        return []
    conn = state_store.conectar(empresa["nit"])
    try:
        return state_store.listar_compras_siigo(conn, desde, hasta, texto)
    finally:
        conn.close()


def _ruta_desde_archivo_origen(archivo_origen: str) -> Path:
    """`compras.archivo_origen` puede haber quedado guardado con separadores
    de Windows ("\\") si la factura se importó en una máquina Windows --
    confirmado real: bases migradas de un PC Windows a un VPS Linux, donde
    Path() trata "\\" como parte del NOMBRE del archivo (no como separador),
    así que la ruta nunca se encuentra aunque el archivo sí exista en disco.
    Normaliza antes de construir el Path -- no importa en qué sistema
    operativo se escribió originalmente ni en cuál se está leyendo ahora."""
    return Path(archivo_origen.replace("\\", "/"))


def obtener_pdf(slug: str, cufe: str) -> bytes | None:
    """Bytes del PDF (representación gráfica) del ZIP o carpeta de origen de
    una factura ya importada, o None si esta no tiene PDF adjunto (o el
    archivo de origen ya no está en disco). Nunca se guarda el PDF en la BD
    -- se lee al vuelo del mismo archivo que dejó `zip_handler` en
    `archivo_origen` (ver docs/05-esquema-datos/modelo-datos.md)."""
    empresa = resolver_empresa(slug)
    db_path = BASE_DATOS_EMPRESAS / f"{empresa['nit']}.db"
    if not db_path.exists():
        return None

    conn = sqlite3.connect(db_path)
    try:
        fila = conn.execute("SELECT archivo_origen FROM compras WHERE cufe = ?", (cufe,)).fetchone()
    finally:
        conn.close()
    if not fila:
        return None

    origen = _ruta_desde_archivo_origen(fila[0])
    if origen.suffix.lower() == ".zip":
        if not origen.exists():
            return None
        with zipfile.ZipFile(origen) as z:
            for nombre in z.namelist():
                if nombre.lower().endswith(".pdf"):
                    return z.read(nombre)
        return None

    candidato_pdf = origen.with_suffix(".pdf")
    return candidato_pdf.read_bytes() if candidato_pdf.exists() else None


def importar_plan_cuentas(slug: str, ruta_excel: str) -> dict:
    """Importa el plan de cuentas exportado de Siigo (Excel). No hay endpoint
    de Siigo para esto -- confirmado 2026-07-21, ver
    docs/04-integracion-siigo/autenticacion-y-endpoints.md -- así que sigue
    siendo un archivo que entrega la empresa.

    Formato confirmado en Fase 0 (docs/05-esquema-datos/plan-cuentas-hielo-super-cool.md):
    6 filas de metadatos arriba, encabezados en la fila 7, columnas en este
    orden exacto: Código, Nombre, Categoría, Clase, Relación con, Maneja
    vencimientos, Diferencia fiscal, Activo, Nivel agrupación.
    """
    import openpyxl

    empresa = resolver_empresa(slug)
    ruta = Path(ruta_excel)
    if not ruta.is_file():
        raise FileNotFoundError(f"El archivo '{ruta}' no existe.")

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active

    cuentas = []
    for fila in ws.iter_rows(min_row=8, values_only=True):
        if not fila or not fila[0]:
            continue
        cuentas.append({
            "codigo": str(fila[0]).strip(),
            "nombre": str(fila[1]).strip() if fila[1] else "",
            "categoria": fila[2],
            "clase": fila[3],
            "relacion_con": fila[4],
            "maneja_vencimientos": fila[5],
            "diferencia_fiscal": fila[6],
            "activo": fila[7],
            "nivel_agrupacion": fila[8],
        })

    if not cuentas:
        raise ValueError(
            "No se encontraron cuentas a partir de la fila 8. ¿Es el formato "
            "esperado (encabezados en fila 7, ver docs/05-esquema-datos/"
            "plan-cuentas-hielo-super-cool.md)?"
        )

    conn = state_store.conectar(empresa["nit"])
    try:
        total = state_store.guardar_plan_cuentas(conn, cuentas)
    finally:
        conn.close()

    transaccionales = sum(1 for c in cuentas if (c["nivel_agrupacion"] or "").strip() == "Transaccional")
    return {"total": total, "transaccionales": transaccionales}


def importar_plan_cuentas_contai(slug: str, ruta_excel: str) -> dict:
    """Importa el plan de cuentas exportado de Contai (Excel) a su propia
    tabla (`plan_cuentas_contai`), separada de la de Siigo -- en Contai la
    información llega directamente como asiento contable (no pasa por un
    módulo de compras que arma el asiento, como sí hace Siigo), así que las
    cuentas traen banderas propias (si reciben movimiento, si requieren
    centro de costo) sin equivalente limpio en el catálogo de Siigo. Se
    conservan las 9 columnas reales tal cual, sin forzarlas a ese formato.

    Formato real confirmado (`contai_plancuentas.xlsx`): encabezados en la
    fila 1, datos desde la fila 2, columnas: Código Cuenta, Nombre Cuenta,
    Tipo de Cuenta, Id. Recibe Movto., Id. Centro Costo, Id. Ajustes,
    Porcentaje Base, Tipo Plazo, Activo."""
    import openpyxl

    empresa = resolver_empresa(slug)
    ruta = Path(ruta_excel)
    if not ruta.is_file():
        raise FileNotFoundError(f"El archivo '{ruta}' no existe.")

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active

    cuentas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or not fila[0]:
            continue
        codigo = str(fila[0]).strip()
        if not codigo:
            continue
        cuentas.append({
            "codigo": codigo,
            "nombre": str(fila[1]).strip() if len(fila) > 1 and fila[1] else "",
            "tipo_cuenta": str(fila[2]).strip() if len(fila) > 2 and fila[2] else None,
            "recibe_movimiento": str(fila[3]).strip() if len(fila) > 3 and fila[3] else None,
            "centro_costo": str(fila[4]).strip() if len(fila) > 4 and fila[4] else None,
            "ajustes": str(fila[5]).strip() if len(fila) > 5 and fila[5] else None,
            "porcentaje_base": fila[6] if len(fila) > 6 else None,
            "tipo_plazo": str(fila[7]).strip() if len(fila) > 7 and fila[7] else None,
            "activo": str(fila[8]).strip() if len(fila) > 8 and fila[8] else None,
        })

    if not cuentas:
        raise ValueError(
            "No se encontraron cuentas a partir de la fila 2. ¿Es el formato "
            "esperado (encabezados en fila 1, columnas de contai_plancuentas.xlsx)?"
        )

    conn = state_store.conectar(empresa["nit"])
    try:
        total = state_store.guardar_plan_cuentas_contai(conn, cuentas)
    finally:
        conn.close()

    transaccionales = sum(1 for c in cuentas if c["recibe_movimiento"] == "S")
    return {"total": total, "transaccionales": transaccionales}


def listar_plan_cuentas_contai(slug: str, solo_transaccionales: bool = False) -> list[dict]:
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        return state_store.listar_plan_cuentas_contai(conn, solo_transaccionales)
    finally:
        conn.close()


def importar_terceros_contai(slug: str, ruta_excel: str) -> dict:
    """Importa el maestro de terceros de Contai (Excel) al caché local
    (`terceros_contai`) -- se usa para detectar, al exportar, qué
    proveedores todavía no existen en Contai y necesitan una fila en el
    plano de terceros nuevos.

    Formato real confirmado (`contai_terceros.xlsx`): encabezados en la
    fila 1 (16 columnas, ver `contai_export.COLUMNAS_TERCERO`). La fila 2
    del archivo real es una fila plantilla/basura del propio exportador de
    Contai (NIT y textos corruptos tipo "Aombre") -- se descarta cualquier
    fila cuyo NIT no tenga al menos un dígito, no se asume que todo dato es
    válido."""
    import openpyxl

    empresa = resolver_empresa(slug)
    ruta = Path(ruta_excel)
    if not ruta.is_file():
        raise FileNotFoundError(f"El archivo '{ruta}' no existe.")

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active
    primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columnas = [str(c).strip() if c else "" for c in primera_fila]

    terceros = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila:
            continue
        nit_crudo = str(fila[0]).strip() if fila[0] else ""
        nit = "".join(ch for ch in nit_crudo if ch.isdigit())
        if not nit:
            continue
        registro = dict(zip(columnas, fila))
        registro["NIT"] = nit
        terceros.append(registro)

    conn = state_store.conectar(empresa["nit"])
    try:
        total = state_store.guardar_terceros_contai(conn, terceros)
    finally:
        conn.close()

    return {"total": total}


def importar_comprobantes_contai(slug: str, ruta_excel: str) -> dict:
    """Importa el catálogo de comprobantes de Contai (Excel) -- solo
    referencia (código + nombre + banderas), pero necesario para poder
    elegir el comprobante de causación desde `config_contai` en vez de
    tenerlo fijo en "00010" (cada empresa/contador puede usar un código
    distinto).

    Formato real confirmado (`contai_comprobantes.xlsx`): encabezados en la
    fila 1 (Comprobante, Nombre, Maneja Consecutivo, Nro. Consecutivo,
    Maneja Codificación, Consecutivo Obligatorio, Vigencia, Documento No
    Obligado, Resolución, Prefijo), datos desde la fila 2."""
    import openpyxl

    empresa = resolver_empresa(slug)
    ruta = Path(ruta_excel)
    if not ruta.is_file():
        raise FileNotFoundError(f"El archivo '{ruta}' no existe.")

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active
    primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columnas = [str(c).strip() if c else "" for c in primera_fila]

    comprobantes = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or not fila[0]:
            continue
        codigo = str(fila[0]).strip()
        if not codigo:
            continue
        registro = dict(zip(columnas, fila))
        registro["Comprobante"] = codigo
        registro["Nombre"] = (str(registro.get("Nombre") or "")).strip()
        comprobantes.append(registro)

    if not comprobantes:
        raise ValueError(
            "No se encontraron comprobantes a partir de la fila 2. ¿Es el formato "
            "esperado (encabezados en fila 1, columnas de contai_comprobantes.xlsx)?"
        )

    conn = state_store.conectar(empresa["nit"])
    try:
        total = state_store.guardar_comprobantes_contai(conn, comprobantes)
    finally:
        conn.close()

    return {"total": total}


def listar_comprobantes_contai(slug: str) -> list[dict]:
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        return state_store.listar_comprobantes_contai(conn)
    finally:
        conn.close()


def importar_movimientos_contai(slug: str, ruta_excel: str) -> dict:
    """Importa el histórico de movimientos ya causados en Contai (Excel) --
    la única fuente que tiene AXON del comportamiento contable real de esta
    empresa en Contai (a diferencia de Siigo, que se consulta en vivo vía
    API). No trae descripción de ítem, solo Cuenta+Tipo+Valor por línea y el
    NIT del proveedor por documento -- se usa para sugerir la cuenta más
    frecuente por proveedor cuando el motor de reglas no resuelve una línea
    (ver `state_store.sugerir_cuenta_historial_contai`).

    Formato real confirmado (`contai_movimientos.xlsx`): encabezados en la
    fila 1 (Cuenta, Comprobante, Fecha(mm/dd/yyyy), Documento, Documento
    Ref., NIT, Detalle, Tipo, Valor, Base, Centro de Costo, Trans. Ext,
    Plazo, Docto Electrónico), datos desde la fila 2. El NIT viene vacío en
    las líneas de crédito (ver contai_export.construir_movimientos) -- se
    guarda tal cual, esas líneas simplemente no participan de la sugerencia
    por proveedor."""
    import openpyxl

    empresa = resolver_empresa(slug)
    ruta = Path(ruta_excel)
    if not ruta.is_file():
        raise FileNotFoundError(f"El archivo '{ruta}' no existe.")

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active

    lineas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or not fila[0]:
            continue
        cuenta = str(fila[0]).strip()
        if not cuenta:
            continue
        nit_crudo = str(fila[5]).strip() if len(fila) > 5 and fila[5] else ""
        nit = "".join(ch for ch in nit_crudo if ch.isdigit())
        lineas.append({
            "proveedor_nit": nit or None,
            "documento": str(fila[3]).strip() if len(fila) > 3 and fila[3] else None,
            "cuenta": cuenta,
            "tipo": int(fila[7]) if len(fila) > 7 and fila[7] is not None else None,
            "valor": fila[8] if len(fila) > 8 else None,
            "fecha": str(fila[2]).strip() if len(fila) > 2 and fila[2] else None,
        })

    if not lineas:
        raise ValueError(
            "No se encontraron líneas a partir de la fila 2. ¿Es el formato "
            "esperado (encabezados en fila 1, columnas de contai_movimientos.xlsx)?"
        )

    conn = state_store.conectar(empresa["nit"])
    try:
        guardado = state_store.guardar_movimientos_contai_historico(conn, lineas)
        resumen = state_store.contar_movimientos_contai_historico(conn)
    finally:
        conn.close()

    return {
        "total": guardado["lineas_insertadas"],  # compatibilidad -- líneas nuevas de ESTA importación
        "lineas_insertadas": guardado["lineas_insertadas"],
        "lineas_omitidas": guardado["lineas_omitidas"],
        "documentos_nuevos": guardado["documentos_nuevos"],
        **resumen,
    }


def obtener_resumen_movimientos_contai(slug: str) -> dict:
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        return state_store.contar_movimientos_contai_historico(conn)
    finally:
        conn.close()


def _parsear_fecha_mmddyyyy(fecha_str: str | None) -> datetime.date | None:
    """El histórico de Contai guarda la fecha tal cual la trae el Excel
    (mm/dd/yyyy, ver importar_movimientos_contai) -- no es comparable como
    texto contra un rango (ni siquiera ordenable: "01/15/2025" < "12/01/2024"
    como string). None si no se puede parsear (no debería pasar con datos
    reales, pero no debe reventar el filtro por fecha)."""
    if not fecha_str:
        return None
    try:
        return datetime.datetime.strptime(fecha_str.strip(), "%m/%d/%Y").date()
    except ValueError:
        return None


def listar_movimientos_contai(
    slug: str, texto: str | None = None, desde: str | None = None, hasta: str | None = None,
) -> list[dict]:
    """Consulta el histórico de movimientos Contai ya importado, agrupado
    por documento (factura) + proveedor -- para responder '¿qué cuenta se
    usó en esta factura / con este proveedor?' sin reabrir el Excel
    original (mismo espíritu que 'Ver en Siigo' / 'Compras en Siigo', ver
    obtener_compra_siigo_de_factura y listar_compras_siigo). `texto` filtra
    por NIT, nombre de proveedor (si ya se importó el maestro de terceros)
    o número de documento. `desde`/`hasta` son fechas ISO (yyyy-mm-dd, lo
    que manda un <input type=date>) y filtran por la fecha del documento."""
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        lineas = state_store.listar_movimientos_contai_historico(conn)
        nombres = state_store.mapa_nombres_terceros_contai(conn)
    finally:
        conn.close()

    # Se agrupa SOLO por documento (número de factura), no por
    # documento+NIT: la línea de crédito (la del medio de pago) viene con
    # proveedor_nit vacío (ver importar_movimientos_contai) -- agrupar
    # también por NIT partiría cada factura en dos filas, una con su
    # proveedor real y otra "fantasma" solo con la línea de crédito.
    grupos: dict[str, dict] = {}
    orden: list[str] = []
    for linea in lineas:
        doc = linea["documento"] or ""
        if doc not in grupos:
            grupos[doc] = {
                "documento": linea["documento"],
                "proveedor_nit": None,
                "proveedor_nombre": None,
                "fecha": linea["fecha"],
                "total_debito": 0.0,
                "lineas": [],
            }
            orden.append(doc)
        grupo = grupos[doc]
        grupo["lineas"].append({"cuenta": linea["cuenta"], "tipo": linea["tipo"], "valor": linea["valor"]})
        if linea["tipo"] == 1:  # débito -- ver contai_export.DEBITO
            grupo["total_debito"] += linea["valor"] or 0.0
        if linea["proveedor_nit"] and not grupo["proveedor_nit"]:
            grupo["proveedor_nit"] = linea["proveedor_nit"]
            grupo["proveedor_nombre"] = nombres.get(linea["proveedor_nit"])

    documentos = [grupos[doc] for doc in orden]

    if texto:
        t = texto.strip().lower()
        documentos = [
            d for d in documentos
            if t in (d["documento"] or "").lower()
            or t in (d["proveedor_nit"] or "").lower()
            or t in (d["proveedor_nombre"] or "").lower()
        ]

    if desde or hasta:
        desde_date = datetime.date.fromisoformat(desde) if desde else None
        hasta_date = datetime.date.fromisoformat(hasta) if hasta else None
        filtrados = []
        for d in documentos:
            fecha_doc = _parsear_fecha_mmddyyyy(d["fecha"])
            if fecha_doc is None:
                continue  # no se puede ubicar en el rango -- se excluye, no se asume
            if desde_date and fecha_doc < desde_date:
                continue
            if hasta_date and fecha_doc > hasta_date:
                continue
            filtrados.append(d)
        documentos = filtrados

    return documentos


def listar_plan_cuentas(slug: str, solo_transaccionales: bool = False) -> list[dict]:
    empresa = resolver_empresa(slug)
    db_path = BASE_DATOS_EMPRESAS / f"{empresa['nit']}.db"
    if not db_path.exists():
        return []
    conn = state_store.conectar(empresa["nit"], base_dir=BASE_DATOS_EMPRESAS)
    try:
        return state_store.listar_plan_cuentas(conn, solo_transaccionales)
    finally:
        conn.close()


def completar_cabecera_faltante_por_empresa(slug: str) -> dict:
    """Rellena tipo_comprobante_id / medio_pago_id en las facturas que
    todavía no los tienen, usando el valor único que ya usa TODA la empresa
    en sus compras ya causadas (ver motor_sugerencias.resolver_cabecera_por_empresa)
    -- caso real: Hielo Super-Cool usa siempre el mismo tipo de comprobante y
    medio de pago, así que las facturas de proveedores nuevos (sin
    preferencia aprendida propia) no tienen por qué quedar vacías. Nunca
    pisa un valor ya presente, y nunca adivina si la empresa tiene más de un
    valor distinto para ese campo -- en ese caso deja el campo como estaba y
    lo reporta como 'no_unico'."""
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        resultado = {}
        for campo in ("tipo_comprobante_id", "medio_pago_id"):
            valor = motor_sugerencias.resolver_cabecera_por_empresa(conn, campo)
            if valor is None:
                valores_distintos = state_store.valores_distintos_cabecera(conn, campo)
                resultado[campo] = {
                    "valor_usado": None,
                    "actualizadas": 0,
                    "motivo": "no_unico" if len(valores_distintos) > 1 else "sin_datos",
                }
                continue
            actualizadas = state_store.completar_cabecera_faltante(conn, campo, valor)
            resultado[campo] = {"valor_usado": valor, "actualizadas": actualizadas, "motivo": None}
        return resultado
    finally:
        conn.close()


def _aplicar_sugerencias(conn: sqlite3.Connection, clasificacion, slug: str | None = None) -> None:
    """Llena lo que `motor_reglas` dejó vacío usando el histórico de Siigo /
    preferencias aprendidas (`motor_sugerencias`), y sube `resuelto_por` a
    `"historico"` cuando la sugerencia fue lo que terminó de completar todas
    las cuentas -- nunca a `"reglas"` (esa marca es solo para una regla de
    negocio confirmada, ver motor_reglas.py). `slug` es opcional (varios
    tests llaman esta función directo sobre un `conn` sin empresa resuelta)
    -- solo se usa para la sugerencia por histórico de Contai, que necesita
    saber el destino de causación y las cuentas de IVA configuradas."""
    tenian_cuenta_antes = all(item.cuenta_contable is not None for item in clasificacion.items)

    # Bug real confirmado en producción (Hielo Super-Cool, factura 2081 de
    # S M BORDADOS Y ESTAMPADOS SAS): bajo la política de IVA no
    # discriminado, las líneas "xml" ya no llevan su propio IVA (se movió
    # al ítem "IVA" aparte, origen="politica_empresa") -- pero el histórico
    # de Siigo puede seguir sugiriendo un iva_tax_id para esa línea (de
    # antes de que la política existiera, o de otro documento del mismo
    # proveedor). Si ese código tiene tarifa > 0% ("IVA Mayor valor de
    # costo" 19%, en este caso), se duplica el IVA: una vez vía el ítem
    # "IVA" y otra vez vía este código inventado en la línea. La política
    # se detecta por la presencia de un ítem origen="politica_empresa" en
    # esta misma factura -- si existe, ninguna línea "xml" recibe iva_tax_id.
    bajo_politica_iva_no_discriminado = any(i.origen == "politica_empresa" for i in clasificacion.items)

    for item in clasificacion.items:
        # tarifa de IVA declarada en el XML para esta línea (tipo exacto
        # "IVA" de TAX_SCHEME_MAP -- no "ReteIVA") -- permite asignar el
        # código del catálogo por porcentaje cuando no hay otra fuente.
        porcentaje_iva_xml = next(
            (imp.get("porcentaje") for imp in item.impuestos if imp.get("tipo") == "IVA"), None,
        )
        sugerido = motor_sugerencias.sugerir_item(
            conn, clasificacion.factura.proveedor_nit, item.descripcion, porcentaje_iva_xml,
        )
        if item.cuenta_contable is None:
            item.cuenta_contable = sugerido["cuenta_contable"]
        # Los ítems inyectados por el motor de reglas (no vienen del XML)
        # nunca llevan IVA ni retención propios:
        # - "otros_impuestos" (impuestos que el parser no supo nombrar):
        #   así lo pidió el usuario explícitamente.
        # - "politica_empresa" (IVA no discriminado de Hielo Super-Cool):
        #   la política dice literalmente "sin impuestos asociados a este
        #   ítem" (docs/02-reglas-negocio/politicas-empresa/901528790-...) --
        #   sin este `continue`, sugerir_item le clavaba "IVA 0%" a todas las
        #   facturas del mes solo porque item.impuestos venía vacío.
        if item.origen in _ORIGENES_INYECTADOS:
            continue
        if item.iva_tax_id is None and not bajo_politica_iva_no_discriminado:
            item.iva_tax_id = sugerido["iva_tax_id"]
        if item.retencion_tax_id is None:
            item.retencion_tax_id = sugerido["retencion_tax_id"]

    # Si el histórico/aprendizaje no resolvió la cuenta de un ítem inyectado
    # ("OTROS IMPUESTOS" o el "IVA" de la política de Hielo Super-Cool), se
    # le asigna la misma cuenta que la(s) línea(s) de GASTO real del mismo
    # documento (origen="xml") -- confirmado por la contadora: la cuenta del
    # IVA no discriminado debe ser la del gasto que acompaña, no una cuenta
    # fija genérica. Solo cuando esas líneas comparten una única cuenta; si
    # el documento mezcla cuentas distintas entre sus líneas, no se adivina
    # cuál usar.
    cuentas_gasto = {
        item.cuenta_contable for item in clasificacion.items
        if item.origen == "xml" and item.cuenta_contable is not None
    }
    if len(cuentas_gasto) == 1:
        for item in clasificacion.items:
            if item.origen in _ORIGENES_INYECTADOS and item.cuenta_contable is None:
                item.cuenta_contable = next(iter(cuentas_gasto))

    # Empresas con destino Contai no tienen histórico de compras vía API
    # (motor_sugerencias.sugerir_item arriba siempre devuelve None para
    # ellas) -- si el usuario ya importó contai_movimientos.xlsx, se usa la
    # cuenta más frecuente de ese proveedor en ese histórico como respaldo.
    # Solo para líneas de gasto real (origen="xml"); los ítems inyectados ya
    # se resuelven arriba heredando de las líneas de gasto del documento.
    if slug is not None and obtener_destino_causacion(slug)["destino_causacion"] == "contai":
        config_contai = obtener_config_contai(slug)
        cuentas_iva_conocidas = set(config_contai.get("cuentas_iva_por_tarifa", {}).values())
        for item in clasificacion.items:
            if item.origen == "xml" and item.cuenta_contable is None:
                sugerida = state_store.sugerir_cuenta_historial_contai(
                    conn, clasificacion.factura.proveedor_nit, cuentas_iva_conocidas,
                )
                if sugerida:
                    item.cuenta_contable = sugerida

    cabecera = motor_sugerencias.sugerir_cabecera(conn, clasificacion.factura.proveedor_nit)
    clasificacion.tipo_comprobante_id = cabecera["tipo_comprobante_id"]
    clasificacion.medio_pago_id = cabecera["medio_pago_id"]

    if not tenian_cuenta_antes and all(item.cuenta_contable is not None for item in clasificacion.items):
        clasificacion.resuelto_por = "historico"


def ejecutar_importar(slug: str, carpeta_relativa: str) -> dict:
    """Corre el pipeline completo (descubrir -> parsear -> clasificar ->
    guardar) sobre data/entrada-dian/<slug>/<carpeta_relativa>. Nunca envía
    nada a Siigo -- eso requiere confirmación explícita aparte (regla 3 de
    CLAUDE.md)."""
    empresa = resolver_empresa(slug)
    nit = empresa["nit"]
    carpeta = ENTRADA_DIAN / slug / carpeta_relativa
    if not carpeta.is_dir():
        raise FileNotFoundError(f"La carpeta '{carpeta}' no existe.")

    conn = state_store.conectar(nit)
    resultado = descubrir_documentos(carpeta)

    nuevas, ya_existentes, con_error_parseo, nit_no_corresponde = 0, 0, 0, 0
    for doc in resultado.documentos:
        if state_store.ya_existe_cufe(conn, doc.cufe):
            ya_existentes += 1
            continue
        try:
            factura = parsear_factura(doc.xml_bytes)
        except ET.ParseError as e:
            state_store.registrar_descartado(
                conn, DocumentoConError(origen=doc.origen, motivo=f"Error al parsear factura: {e}")
            )
            con_error_parseo += 1
            continue
        # Si el XML trae el NIT del receptor (comprador) y no coincide con el
        # de esta empresa, no se importa -- protege contra mezclar facturas
        # de otra empresa por una carpeta o subida equivocada (bug real
        # posible desde que existe la subida manual en modo SaaS, ver
        # docs/06-multiempresa-saas/aislamiento-datos.md). Si el XML no trae
        # ese dato (factura.receptor_nit es None) se deja pasar -- no se
        # puede validar lo que no viene, y varios proveedores/fixtures no lo
        # incluyen.
        if factura.receptor_nit and factura.receptor_nit != nit:
            state_store.registrar_descartado(
                conn, DocumentoConError(
                    origen=doc.origen,
                    motivo=(
                        f"CUFE {doc.cufe}: la factura es para el NIT {factura.receptor_nit}, "
                        f"no para el NIT de esta empresa ({nit}) -- no se importó para no "
                        "mezclar datos entre empresas."
                    ),
                ),
            )
            nit_no_corresponde += 1
            continue
        _marcar_perfil_fiscal_automatico(factura.proveedor_nit, factura.proveedor_nombre, factura.responsabilidades_fiscales)
        clasificacion = clasificar_factura(factura, nit_empresa=nit)
        _aplicar_sugerencias(conn, clasificacion, slug=slug)
        state_store.guardar_resultado(conn, clasificacion, archivo_origen=doc.origen)
        nuevas += 1

    for dup in resultado.duplicados:
        state_store.registrar_descartado(conn, dup)
    for err in resultado.con_error:
        state_store.registrar_descartado(conn, err)
    for no_fact in resultado.no_facturas:
        state_store.registrar_descartado(conn, no_fact)

    return {
        "empresa": empresa["nombre"],
        "nit": nit,
        "carpeta": str(carpeta),
        "nuevas": nuevas,
        "ya_existentes": ya_existentes,
        "duplicados": len(resultado.duplicados),
        "no_facturas": len(resultado.no_facturas),
        "con_error": len(resultado.con_error) + con_error_parseo,
        "nit_no_corresponde": nit_no_corresponde,
    }


_CAMPOS_CABECERA_APRENDIBLES = {"tipo_comprobante_id", "medio_pago_id"}


def actualizar_factura(slug: str, cufe: str, campos: dict) -> dict:
    """Actualiza campos de cabecera de una factura ya importada (tipo de
    comprobante, medio de pago, modo de pago Contai) desde el panel de
    detalle. Solo tipo_comprobante_id/medio_pago_id se aprenden para la
    próxima importación de ese proveedor (ver motor_sugerencias.aprender) --
    modo_pago_contai queda fuera a propósito: es una decisión por factura
    (dos facturas del mismo proveedor pueden pagarse distinto), no un patrón
    de clasificación del proveedor como los otros dos."""
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        fila = conn.execute("SELECT proveedor_nit FROM compras WHERE cufe = ?", (cufe,)).fetchone()
        if fila is None:
            raise ValueError(f"No existe una factura con CUFE '{cufe}' en esta empresa.")
        proveedor_nit = fila[0]
        state_store.actualizar_compra_campos(conn, cufe, campos)
        for campo, valor in campos.items():
            if campo in _CAMPOS_CABECERA_APRENDIBLES:
                motor_sugerencias.aprender(conn, campo, proveedor_nit, None, valor)
        return {"cufe": cufe, **campos}
    finally:
        conn.close()


def actualizar_item(slug: str, cufe: str, item_id: int, campos: dict) -> dict:
    """Actualiza cuenta contable / IVA / retefuente de una línea ya
    importada, y aprende la elección por proveedor+descripción del ítem."""
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        item = state_store.obtener_detalle(conn, item_id)
        if item is None or item["cufe"] != cufe:
            raise ValueError(f"El ítem {item_id} no pertenece a la factura '{cufe}' de esta empresa.")
        state_store.actualizar_detalle_campos(conn, item_id, campos)
        for campo, valor in campos.items():
            motor_sugerencias.aprender(conn, campo, item["proveedor_nit"], item["descripcion"], valor)
        return {"id": item_id, **campos}
    finally:
        conn.close()


def replicar_campo_item(slug: str, cufe: str, item_id: int, campo: str) -> dict:
    """Replica el valor actual de `campo` (cuenta_contable | iva_tax_id |
    retencion_tax_id) de un ítem a todos los demás ítems de la misma
    factura -- "modifica una línea y replícalo a las demás"."""
    if campo not in CAMPOS_ITEM_REPLICABLES:
        raise ValueError(f"Campo no replicable: '{campo}'. Válidos: {', '.join(CAMPOS_ITEM_REPLICABLES)}.")
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        origen = state_store.obtener_detalle(conn, item_id)
        if origen is None or origen["cufe"] != cufe:
            raise ValueError(f"El ítem {item_id} no pertenece a la factura '{cufe}' de esta empresa.")
        valor = origen[campo]
        afectados = 0
        for item in state_store.listar_detalle_por_compra(conn, origen["compra_id"]):
            if item["id"] == item_id:
                continue
            state_store.actualizar_detalle_campos(conn, item["id"], {campo: valor})
            motor_sugerencias.aprender(conn, campo, item["proveedor_nit"], item["descripcion"], valor)
            afectados += 1
        return {"campo": campo, "valor": valor, "lineas_actualizadas": afectados}
    finally:
        conn.close()


def buscar_candidatos_recalculo(slug: str, cufe: str, item_id: int, campo: str, desde: str, hasta: str) -> dict:
    """Previsualización de "recalcular sin reimportar": busca ítems de OTRAS
    facturas del mismo proveedor, dentro del rango de fechas activo en la
    bandeja, con una descripción parecida (no necesariamente idéntica -- ver
    motor_sugerencias.descripciones_similares) a la del ítem que se acaba de
    corregir. No modifica nada todavía -- el usuario confirma cuáles de
    estos candidatos sí quiere actualizar en `aplicar_recalculo_masivo`.
    `desde`/`hasta` son obligatorios por la misma razón que en
    validar_completitud: nunca se recalcula contra todo el historial."""
    if campo not in CAMPOS_ITEM_REPLICABLES:
        raise ValueError(f"Campo no soportado para recalcular: '{campo}'. Válidos: {', '.join(CAMPOS_ITEM_REPLICABLES)}.")
    if not desde or not hasta:
        raise ValueError(
            "Selecciona primero un rango de fechas (Desde/Hasta) en la bandeja -- "
            "recalcular nunca se corre contra toda la tabla."
        )

    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        origen = state_store.obtener_detalle(conn, item_id)
        if origen is None or origen["cufe"] != cufe:
            raise ValueError(f"El ítem {item_id} no pertenece a la factura '{cufe}' de esta empresa.")
        valor = origen[campo]
        if valor is None:
            raise ValueError("Este campo está vacío en el ítem de origen -- no hay nada que propagar.")

        candidatos = []
        for item in state_store.listar_items_por_proveedor_y_rango(conn, origen["proveedor_nit"], desde, hasta):
            if item["id"] == item_id or item[campo] == valor:
                continue
            similitud = motor_sugerencias.similitud_descripcion(item["descripcion"], origen["descripcion"])
            if similitud < motor_sugerencias.UMBRAL_SIMILITUD_DESCRIPCION:
                continue
            candidatos.append({**item, "similitud": round(similitud, 2)})
        candidatos.sort(key=lambda c: c["similitud"], reverse=True)

        return {
            "campo": campo,
            "valor": valor,
            "descripcion_origen": origen["descripcion"],
            "candidatos": candidatos,
        }
    finally:
        conn.close()


def aplicar_recalculo_masivo(slug: str, campo: str, valor: str, item_ids: list[int]) -> dict:
    """Aplica `campo = valor` a exactamente los ítems que el usuario
    confirmó en `buscar_candidatos_recalculo` -- nunca a ciegas a todo lo
    que encontró el matcheo difuso. También aprende cada uno (proveedor +
    su propia descripción), igual que una edición manual normal."""
    if campo not in CAMPOS_ITEM_REPLICABLES:
        raise ValueError(f"Campo no soportado para recalcular: '{campo}'. Válidos: {', '.join(CAMPOS_ITEM_REPLICABLES)}.")
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        actualizados = 0
        for item_id in item_ids:
            item = state_store.obtener_detalle(conn, item_id)
            if item is None:
                continue
            state_store.actualizar_detalle_campos(conn, item_id, {campo: valor})
            motor_sugerencias.aprender(conn, campo, item["proveedor_nit"], item["descripcion"], valor)
            actualizados += 1
        return {"actualizados": actualizados}
    finally:
        conn.close()


def eliminar_facturas(slug: str, cufes: list[str]) -> dict:
    """Borra permanentemente facturas importadas por error -- el frontend
    debe confirmar antes de llamar esto, no hay deshacer (ver
    state_store.eliminar_compras)."""
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        total = state_store.eliminar_compras(conn, cufes)
        return {"eliminadas": total}
    finally:
        conn.close()


def _ruta_config_proveedor(nit_proveedor: str) -> Path:
    return CONFIG_PROVEEDORES_DIR / f"{nit_proveedor}.json"


def _leer_config_proveedor_raw(nit_proveedor: str, nombre_proveedor: str) -> dict:
    ruta = _ruta_config_proveedor(nit_proveedor)
    if ruta.exists():
        with open(ruta, encoding="utf-8") as f:
            return json.load(f)
    return {"nit": nit_proveedor, "nombre": nombre_proveedor, "slug": "", "descripcion_md": "", "comportamiento": {}}


def _guardar_config_proveedor_raw(nit_proveedor: str, config: dict) -> None:
    ruta = _ruta_config_proveedor(nit_proveedor)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def marcar_proveedor_autorretenedor(nit_proveedor: str, nombre_proveedor: str, autorretenedor: bool) -> dict:
    """Marca/desmarca un proveedor como autorretenedor -- motor_sugerencias
    nunca sugiere retención para un proveedor marcado así, porque la
    autorretención la asume el proveedor, no el comprador. Escribe en
    config/proveedores/<nit>.json (comportamiento.autorretenedor), que SÍ se
    versiona en git (a diferencia de config/empresas/) y es global por NIT
    de proveedor, no por empresa-cliente. Es la primera vez que el "perfil
    de proveedor" que motor_reglas.cargar_config_proveedor ya leía hace algo
    funcional -- antes solo se anotaba en las notas de la factura."""
    config = _leer_config_proveedor_raw(nit_proveedor, nombre_proveedor)
    config.setdefault("comportamiento", {})
    config["comportamiento"]["autorretenedor"] = autorretenedor
    _guardar_config_proveedor_raw(nit_proveedor, config)
    return {"nit": nit_proveedor, "autorretenedor": autorretenedor}


# Códigos DIAN de responsabilidad fiscal (cac:PartyTaxScheme/cbc:TaxLevelCode)
# que activan el marcado automático del perfil de proveedor al importar --
# confirmado con QUALA ("O-13;O-15;O-23"), KOPPS ("O-13;O-15") y COMMERK
# ("O-13") en facturas reales de julio de Distribuidora El Manantial.
_CODIGO_AUTORRETENEDOR = "O-15"
_CODIGO_GRAN_CONTRIBUYENTE = "O-13"


def _marcar_perfil_fiscal_automatico(nit_proveedor: str, nombre_proveedor: str, responsabilidades: list[str]) -> None:
    """Si el propio XML del proveedor declara que es autorretenedor y/o gran
    contribuyente, lo marca en config/proveedores/<nit>.json sin esperar a
    que alguien lo haga a mano desde la bandeja -- así motor_sugerencias.
    es_autorretenedor() ya lo sabe desde la primera factura que se importa de
    ese proveedor. Nunca DESMARCA nada: que una factura puntual no traiga el
    código no prueba que el proveedor haya dejado de serlo (podría ser un
    XML incompleto, no un cambio real de condición fiscal)."""
    cambios = {}
    if _CODIGO_AUTORRETENEDOR in responsabilidades:
        cambios["autorretenedor"] = True
    if _CODIGO_GRAN_CONTRIBUYENTE in responsabilidades:
        cambios["gran_contribuyente"] = True
    if not cambios:
        return
    config = _leer_config_proveedor_raw(nit_proveedor, nombre_proveedor)
    config.setdefault("comportamiento", {})
    config["comportamiento"].update(cambios)
    _guardar_config_proveedor_raw(nit_proveedor, config)


def _normalizar_encabezado(texto) -> str:
    """Quita tildes/caracteres no ASCII antes de comparar nombres de columna.
    Necesario porque el listado real de la DIAN (confirmado contra un
    archivo de Hielo Super-Cool) trae 'Fecha Emisión' con el carácter de
    reemplazo U+FFFD en vez de 'ó' -- un problema de codificación del propio
    exportador, no nuestro. Comparar por prefijo ASCII estable en vez de
    igualdad exacta evita que ese tipo de corrupción rompa el parseo."""
    if texto is None:
        return ""
    descompuesto = unicodedata.normalize("NFKD", str(texto))
    return descompuesto.encode("ascii", "ignore").decode("ascii").strip().lower()


def _indice_columnas_listado(encabezados: list) -> dict[str, int]:
    indice: dict[str, int] = {}
    for i, encabezado in enumerate(encabezados):
        n = _normalizar_encabezado(encabezado)
        if n == "tipo de documento":
            indice["tipo_documento"] = i
        elif n == "cufe/cude":
            indice["cufe"] = i
        elif n.startswith("fecha emisi"):
            indice["fecha_emision"] = i
        elif n == "nit receptor":
            indice["nit_receptor"] = i
        elif n == "nit emisor":
            indice["proveedor_nit"] = i
        elif n == "nombre emisor":
            indice["proveedor_nombre"] = i
        elif n == "folio":
            indice["folio"] = i
        elif n == "prefijo":
            indice["prefijo"] = i
        elif n == "total":
            indice["total"] = i
        elif n == "grupo":
            indice["grupo"] = i
    return indice


_COLUMNAS_LISTADO_REQUERIDAS = ("tipo_documento", "cufe", "fecha_emision", "nit_receptor", "grupo")


def _fecha_listado_a_iso(valor) -> str | None:
    """El listado real trae 'Fecha Emisión' como texto 'DD-MM-YYYY' (no ISO,
    confirmado contra archivo real) -- pero se maneja también el caso de que
    Excel la haya guardado como fecha nativa, por robustez."""
    if valor is None:
        return None
    if isinstance(valor, datetime.datetime):
        return valor.date().isoformat()
    if isinstance(valor, datetime.date):
        return valor.isoformat()
    try:
        return datetime.datetime.strptime(str(valor).strip(), "%d-%m-%Y").date().isoformat()
    except ValueError:
        return None


def validar_completitud(slug: str, carpeta_relativa: str, nombre_archivo: str, desde: str, hasta: str) -> dict:
    """Compara el listado de compras que exporta el portal de la DIAN contra
    lo que YA está importado en la bandeja de esta empresa -- no contra el
    ZIP crudo (ver docs/03-ingesta-dian/validador-completitud.md para el
    formato del listado, confirmado contra un archivo real). Requiere que
    `nombre_archivo` ya esté en data/entrada-dian/<slug>/<carpeta_relativa>/
    -- si quien valida no tiene acceso al disco del servidor (SaaS), ver
    `validar_completitud_archivo_subido`."""
    ruta = ENTRADA_DIAN / slug / carpeta_relativa / nombre_archivo
    if not ruta.is_file():
        raise FileNotFoundError(f"El archivo '{ruta}' no existe.")
    return _validar_completitud_desde_ruta(slug, ruta, desde, hasta)


def validar_completitud_archivo_subido(slug: str, nombre_archivo: str, contenido: bytes, desde: str, hasta: str) -> dict:
    """Como `validar_completitud`, pero para cuando el listado se sube desde
    el navegador en vez de ya estar en el servidor -- mismo caso de uso que
    `guardar_archivos_subidos`: en el modo SaaS quien valida no tiene ni
    sabe de una ruta local. El archivo se guarda de todas formas (no es
    efímero) para poder revisarlo después sin tener que volver a subirlo."""
    resolver_empresa(slug)
    nombre_seguro = Path(nombre_archivo).name or "listado.xlsx"
    carpeta = ENTRADA_DIAN / slug / "_listados_subidos"
    carpeta.mkdir(parents=True, exist_ok=True)
    marca = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    ruta = carpeta / f"{marca}_{nombre_seguro}"
    ruta.write_bytes(contenido)
    return _validar_completitud_desde_ruta(slug, ruta, desde, hasta)


def _validar_completitud_desde_ruta(slug: str, ruta: Path, desde: str, hasta: str) -> dict:
    """Núcleo compartido por `validar_completitud` y
    `validar_completitud_archivo_subido` -- la única diferencia entre las
    dos es de dónde sale `ruta` (ya en el servidor, o recién subida).

    La comparación SIEMPRE se acota al rango `desde`/`hasta` que el usuario
    tiene activo en la bandeja -- son obligatorios, nunca se compara contra
    toda la tabla histórica de la empresa (sería lento y un listado de un
    mes puntual generaría falsos "faltantes" contra facturas de otros
    meses)."""
    import openpyxl

    if not desde or not hasta:
        raise ValueError(
            "Selecciona primero un rango de fechas (Desde/Hasta) en la bandeja -- "
            "la validación de completitud nunca se corre contra toda la tabla."
        )

    empresa = resolver_empresa(slug)

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active
    primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    indice = _indice_columnas_listado(primera_fila)
    faltan = [c for c in _COLUMNAS_LISTADO_REQUERIDAS if c not in indice]
    if faltan:
        raise ValueError(
            f"El archivo no tiene el formato esperado del listado DIAN -- no se encontraron "
            f"las columnas: {', '.join(faltan)}."
        )

    def _col(fila, clave):
        i = indice.get(clave)
        return fila[i] if i is not None and i < len(fila) else None

    filas_listado = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or not _col(fila, "cufe"):
            continue
        if _normalizar_encabezado(_col(fila, "grupo")) != "recibido":
            continue
        if str(_col(fila, "nit_receptor") or "").strip() != empresa["nit"]:
            continue
        if _normalizar_encabezado(_col(fila, "tipo_documento")) != "factura electronica":
            continue
        fecha_iso = _fecha_listado_a_iso(_col(fila, "fecha_emision"))
        if not fecha_iso or fecha_iso < desde or fecha_iso > hasta:
            continue
        filas_listado.append({
            "cufe": str(_col(fila, "cufe")).strip(),
            "fecha": fecha_iso,
            "folio": _col(fila, "folio"),
            "prefijo": _col(fila, "prefijo"),
            "proveedor_nit": _col(fila, "proveedor_nit"),
            "proveedor_nombre": _col(fila, "proveedor_nombre"),
            "total": _col(fila, "total"),
        })

    if not filas_listado:
        raise ValueError(
            f"El listado no trae ninguna factura recibida por esta empresa entre {desde} y {hasta} -- "
            "revisa que sea el archivo correcto o que el rango incluya ese periodo."
        )

    conn = state_store.conectar(empresa["nit"])
    try:
        cufes_bandeja = {
            fila[0] for fila in conn.execute(
                "SELECT cufe FROM compras WHERE fecha_emision >= ? AND fecha_emision <= ?", (desde, hasta)
            ).fetchall()
        }
    finally:
        conn.close()

    cufes_listado = {f["cufe"] for f in filas_listado}
    faltantes = [f for f in filas_listado if f["cufe"] not in cufes_bandeja]
    faltantes.sort(key=lambda f: (f["fecha"], f["folio"] or 0))

    return {
        "rango": {"desde": desde, "hasta": hasta},
        "total_listado": len(filas_listado),
        "total_bandeja_en_rango": len(cufes_bandeja),
        "faltantes": faltantes,
        "sobrantes_en_bandeja": len(cufes_bandeja - cufes_listado),
    }


def reporte_faltantes_completitud_xlsx(faltantes: list[dict]) -> bytes:
    """Arma un .xlsx descargable con las facturas que `validar_completitud`
    marcó como faltantes -- exactamente las mismas que ya se le mostraron al
    usuario en la tabla (no vuelve a comparar nada contra la bandeja), para
    que las pueda compartir o gestionar aparte. Pedido explícito del usuario
    (agosto 2026): número de factura, fecha, NIT y nombre del emisor, y
    valor -- en ese orden (NIT/nombre van antes del valor)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faltantes"
    ws.append(["Número de factura", "Fecha", "NIT emisor", "Nombre emisor", "Valor"])
    for fila in faltantes:
        numero = f"{fila.get('prefijo') or ''}{fila.get('folio') or ''}"
        fecha = fila.get("fecha")
        celda_fecha = datetime.date.fromisoformat(fecha) if fecha else None
        ws.append([numero, celda_fecha, fila.get("proveedor_nit"), fila.get("proveedor_nombre"), fila.get("total")])

    for fila_excel in range(2, ws.max_row + 1):
        ws.cell(row=fila_excel, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=fila_excel, column=5).number_format = "#,##0"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _factura_proveedor_de(factura: dict) -> str:
    """Prefijo+número del proveedor concatenados, en el mismo formato en que
    el caché `compras_siigo` guarda `factura_proveedor` (ver
    `_mapear_compra_siigo`) -- la llave del cruce antidúplicados.

    Bug real confirmado (factura 2081 de S M BORDADOS Y ESTAMPADOS SAS,
    Hielo Super-Cool): un folio DIAN sin prefijo se envía con
    `siigo_payload.PREFIJO_RESPALDO` ("FC", ver ese módulo) -- si acá se
    reconstruía con prefijo vacío en vez de ese mismo respaldo, la llave no
    coincidía con la que de verdad quedó en Siigo, y tanto el
    antidúplicados como "Ver en Siigo" fallaban en silencio para cualquier
    factura con folio puramente numérico."""
    prefijo = factura.get("prefijo") or siigo_payload.PREFIJO_RESPALDO
    numero = factura.get("numero_puro") or ""
    return f"{prefijo}{numero}" if numero else (factura.get("numero_factura") or "")


def _motivo_duplicado(conn: sqlite3.Connection, factura: dict) -> str | None:
    """Protección antidúplicados: causar dos veces la misma factura del
    proveedor en Siigo es un error contable real. Dos fuentes, en orden:
    1. `estado_siigo == 'enviado'`: ya se envió desde este mismo sistema.
    2. El caché local de compras causadas (`compras_siigo`): la factura ya
       existe en Siigo, causada por otra vía (el aplicativo anterior, o a
       mano en Siigo Nube). El caché se alimenta del botón "Descargar" en
       la pantalla Compras en Siigo Y de cada envío exitoso nuestro -- pero
       si nunca se ha descargado para esta empresa, esta segunda fuente no
       ve nada: conviene descargar antes de un lote grande."""
    if factura.get("estado_siigo") == "enviado":
        siigo_id = factura.get("siigo_id") or "sin id"
        return f"Ya fue enviada a Siigo desde aquí (id {siigo_id})."
    factura_proveedor = _factura_proveedor_de(factura)
    if factura_proveedor and state_store.existe_compra_siigo(conn, factura["proveedor_nit"], factura_proveedor):
        return (
            f"Ya existe en Siigo una compra de este proveedor con la factura {factura_proveedor} "
            "(según el caché de 'Compras en Siigo') -- enviarla de nuevo la duplicaría."
        )
    return None


_RESPONSABILIDADES_FISCALES_SIIGO = {"R-99-PN", "O-13", "O-15", "O-23", "O-47"}


def _extraer_tercero_de_origen(archivo_origen: str, cufe: str) -> dict | None:
    """Relee el XML original de la factura (ZIP o XML suelto en
    data/entrada-dian/, ver `compras.archivo_origen`) y saca los datos del
    emisor -- la razón por la que nunca se guarda el XML crudo en la base:
    `archivo_origen` + `cufe` bastan para volver a él cuando hace falta
    (docstring de state_store)."""
    ruta = _ruta_desde_archivo_origen(archivo_origen)
    if not ruta.is_file():
        return None
    candidatos: list[bytes] = []
    if ruta.suffix.lower() == ".zip":
        with zipfile.ZipFile(ruta) as z:
            candidatos = [z.read(n) for n in z.namelist() if n.lower().endswith(".xml")]
    else:
        candidatos = [ruta.read_bytes()]
    for xml_bytes in candidatos:
        try:
            if extraer_cufe(xml_bytes) == cufe:
                return extraer_tercero(xml_bytes)
        except ET.ParseError:
            continue
    return None


def _payload_tercero(tercero: dict) -> dict:
    """Arma el body de POST /v1/customers desde los datos del XML DIAN.
    `person_type`/`name`: para NIT (id_type 31) es una empresa con un solo
    campo de nombre; para cédula, Siigo exige nombre y apellido separados --
    se parte por el primer espacio (mejor aproximación posible desde el XML,
    que trae la razón social como un solo texto)."""
    id_type = tercero.get("id_type") or "31"
    es_empresa = id_type == "31"
    nombre = (tercero.get("nombre") or "SIN NOMBRE").strip()
    if es_empresa:
        name = [nombre[:100]]
    else:
        partes = nombre.split(" ", 1)
        name = [partes[0][:100], (partes[1] if len(partes) > 1 else partes[0])[:100]]

    # El XML DIAN puede traer varios códigos de responsabilidad fiscal juntos
    # en un solo TaxLevelCode, separados por ";" (confirmado real: QUALA trae
    # "O-13;O-15;O-23") -- compararlo completo contra _RESPONSABILIDADES_FISCALES_SIIGO
    # nunca hacía match y siempre caía a R-99-PN, aunque el proveedor sí
    # tuviera un código válido. Hay que partirlo primero.
    codigos = [c.strip() for c in (tercero.get("tax_level_code") or "").split(";") if c.strip()]
    codigos_validos = [c for c in codigos if c in _RESPONSABILIDADES_FISCALES_SIIGO]
    fiscal_responsibilities = [{"code": c} for c in codigos_validos] or [{"code": "R-99-PN"}]

    payload = {
        "type": "Supplier",
        "person_type": "Company" if es_empresa else "Person",
        "id_type": id_type,
        "identification": tercero["nit"],
        "name": name,
        "active": True,
        "fiscal_responsibilities": fiscal_responsibilities,
        "address": {
            "address": (tercero.get("direccion") or "NO INFORMADA")[:256],
            "city": {
                "country_code": "Co",
                "state_code": tercero.get("departamento_codigo") or "",
                "city_code": tercero.get("ciudad_codigo") or "",
            },
        },
        "contacts": [{"first_name": nombre[:50]}],
    }
    if tercero.get("digito_verificacion"):
        payload["check_digit"] = tercero["digito_verificacion"]
    telefono_digitos = re.sub(r"\D", "", tercero.get("telefono") or "")
    if telefono_digitos:
        # Confirmado con envíos reales:
        # - COMMERK (NIT 800007955): Siigo rechaza "phones[0].number" con
        #   guiones u otros caracteres no numéricos ("Invalid data type:
        #   number") -- el XML DIAN sí trae teléfonos con guion (ej.
        #   "322-3677140"), así que hay que limpiarlos antes de enviarlos.
        # - SYS FMQ (NIT 901079686): Siigo además rechaza más de 10 dígitos
        #   ("length_max") -- el XML a veces trae el indicativo de país
        #   ("+573223047049", 12 dígitos). Los últimos 10 dígitos son el
        #   número colombiano real sin importar si el XML incluyó o no el
        #   indicativo -- se toman esos, no los primeros N.
        payload["phones"] = [{"number": telefono_digitos[-10:]}]
    if tercero.get("correo"):
        payload["contacts"][0]["email"] = tercero["correo"]
    return payload


def _asegurar_proveedor_en_siigo(conn: sqlite3.Connection, token: str, partner_id: str, factura: dict) -> str | None:
    """Si el NIT del proveedor no existe como tercero en Siigo, lo crea con
    los datos del XML original (autorizado por el usuario: 'si algún nit no
    existe en siigo, créalo antes'). Devuelve un mensaje de error si no se
    pudo asegurar (la factura no debe enviarse en ese caso), o None si el
    proveedor ya existe / quedó creado."""
    nit = factura["proveedor_nit"]
    if siigo_client.obtener_nombre_proveedor(token, partner_id, nit):
        return None

    fila = conn.execute("SELECT archivo_origen FROM compras WHERE cufe = ?", (factura["cufe"],)).fetchone()
    archivo_origen = fila[0] if fila else None
    tercero = _extraer_tercero_de_origen(archivo_origen, factura["cufe"]) if archivo_origen else None
    if not tercero or not tercero.get("nit"):
        return (
            f"El proveedor NIT {nit} no existe en Siigo y no se pudieron releer sus datos "
            f"del XML original ({archivo_origen or 'sin archivo de origen'}) para crearlo."
        )
    if not tercero.get("ciudad_codigo") or not tercero.get("departamento_codigo"):
        return (
            f"El proveedor NIT {nit} no existe en Siigo y el XML no trae los códigos DANE de "
            "ciudad/departamento que Siigo exige para crearlo -- créalo a mano en Siigo Nube primero."
        )
    try:
        siigo_client.crear_customer(token, partner_id, _payload_tercero(tercero))
        state_store.guardar_nombre_proveedor_siigo(conn, nit, tercero["nombre"])
        return None
    except siigo_client.SiigoError as e:
        return f"No se pudo crear el proveedor NIT {nit} en Siigo: {e}"


def obtener_compra_siigo_de_factura(slug: str, cufe: str) -> dict | None:
    """Cómo quedó causada esta factura en Siigo, según el caché local de
    'Compras en Siigo' -- botón 'Ver en Siigo' del panel de detalle. Cruza
    por proveedor + prefijo+número (la misma llave del antidúplicados), no
    por CUFE, porque Siigo no conoce el CUFE. `None` si no está en el caché
    -- que puede significar que no está causada O que ese periodo no se ha
    descargado; el frontend lo aclara, acá no se adivina."""
    empresa = resolver_empresa(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        fila = conn.execute(
            "SELECT proveedor_nit, prefijo, numero_puro, numero_factura FROM compras WHERE cufe = ?", (cufe,)
        ).fetchone()
        if fila is None:
            raise ValueError(f"No existe una factura con CUFE '{cufe}' en esta empresa.")
        proveedor_nit, prefijo, numero_puro, numero_factura = fila
        factura_proveedor = _factura_proveedor_de({
            "prefijo": prefijo, "numero_puro": numero_puro, "numero_factura": numero_factura,
        })
        return state_store.obtener_compra_siigo(conn, proveedor_nit, factura_proveedor)
    finally:
        conn.close()


def corregir_iva_duplicado_enviadas(slug: str, cufes: list[str]) -> dict:
    """Borra y vuelve a causar en Siigo compras YA enviadas cuyo ítem de IVA
    (política 'iva_no_discriminado') se guardó duplicado por un bug ya
    corregido en motor_reglas -- uso EXCLUSIVO de esta corrección puntual,
    confirmada por el usuario factura por factura antes de tocar nada (no es
    un reintento normal: de verdad borra un registro contable real).

    Requiere que el valor local (`detalle_compras.valor_unitario` del ítem
    'politica_empresa') ya haya sido corregido antes de llamar esto -- este
    método solo sincroniza Siigo con lo que ya dice la base local, nunca
    recalcula nada por su cuenta.

    Orden por factura, pensado para minimizar la ventana de riesgo: primero
    arma y valida el payload NUEVO (correcto) -- si eso falla, no se borra
    nada, la compra vieja (incorrecta) sigue intacta en Siigo, seguro de
    reintentar. Solo si el payload nuevo es válido se borra la compra vieja
    y se crea la nueva. Si el borrado funciona pero la creación falla, la
    factura queda marcada 'error' con el detalle completo -- ese hueco
    (borrada en Siigo pero no recreada) requiere atención inmediata, nunca
    se reintenta solo."""
    empresa = resolver_empresa(slug)
    conexion = obtener_conexion_siigo(slug)
    if not conexion["configurado"]:
        raise ValueError(
            "Esta empresa todavía no tiene usuario/access_key configurados -- "
            "complétalos primero en el menú 'Conexión Siigo'."
        )
    token = siigo_client.autenticar(conexion["usuario"], conexion["access_key"])
    partner_id = conexion["partner_id"]

    facturas_por_cufe = {f["cufe"]: f for f in listar_facturas(slug)}
    conn = state_store.conectar(empresa["nit"])
    try:
        catalogo_taxes = state_store.listar_catalogo_siigo(conn, "taxes")
        corregidas, con_error = 0, 0
        detalle = []
        for cufe in cufes:
            factura = facturas_por_cufe.get(cufe)
            if factura is None:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "No existe esa factura en esta empresa."})
                continue
            if factura.get("estado_siigo") != "enviado" or not factura.get("siigo_id"):
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "Esta factura no está marcada como enviada -- usa confirmar_envio_siigo, no esta corrección."})
                continue

            siigo_id_viejo = factura["siigo_id"]
            armado = siigo_payload.construir_payload(factura, catalogo_taxes)
            if armado["motivos_bloqueo"]:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "; ".join(armado["motivos_bloqueo"])})
                continue

            try:
                siigo_client.eliminar_purchase(token, partner_id, siigo_id_viejo)
            except siigo_client.SiigoError as e:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": f"No se pudo borrar la compra vieja ({siigo_id_viejo}), no se tocó nada más: {e}"})
                continue

            try:
                respuesta = siigo_client.crear_purchase(token, partner_id, armado["payload"])
                siigo_id_nuevo = str(respuesta.get("id", ""))
                state_store.registrar_resultado_envio_siigo(conn, cufe, "enviado", siigo_id=siigo_id_nuevo, siigo_error=None)
                try:
                    fila_cache = _mapear_compra_siigo(respuesta, factura["proveedor_nombre"])
                    state_store.guardar_compras_siigo(conn, [fila_cache], reemplazar_todo=False)
                except (KeyError, TypeError, ValueError):
                    pass
                corregidas += 1
                detalle.append({"cufe": cufe, "ok": True, "siigo_id_viejo": siigo_id_viejo, "siigo_id_nuevo": siigo_id_nuevo})
            except siigo_client.SiigoError as e:
                error_json = json.dumps(
                    {"error": f"BORRADA en Siigo (id {siigo_id_viejo}) pero NO se pudo recrear -- requiere atención manual inmediata: {e}",
                     "payload_enviado": armado["payload"]},
                    ensure_ascii=False,
                )
                state_store.registrar_resultado_envio_siigo(conn, cufe, "error", siigo_id=None, siigo_error=error_json)
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": f"BORRADA (id {siigo_id_viejo}) pero no se pudo recrear: {e}"})
        return {"corregidas": corregidas, "con_error": con_error, "detalle": detalle}
    finally:
        conn.close()


def previsualizar_envio_siigo(slug: str, cufes: list[str]) -> list[dict]:
    """Arma, SIN tocar la red, el payload exacto que se enviaría a Siigo por
    cada cufe pedido -- para mostrarlo en el modal de confirmación antes de
    que el usuario apriete "Confirmar y enviar" (CLAUDE.md regla 3: nunca se
    envía nada sin confirmación explícita). Nunca autentica ni llama a
    Siigo -- se puede invocar tantas veces como haga falta para refrescar
    la vista previa."""
    empresa = resolver_empresa(slug)
    facturas_por_cufe = {f["cufe"]: f for f in listar_facturas(slug)}
    conn = state_store.conectar(empresa["nit"])
    try:
        catalogo_taxes = state_store.listar_catalogo_siigo(conn, "taxes")

        resultado = []
        for cufe in cufes:
            factura = facturas_por_cufe.get(cufe)
            if factura is None:
                resultado.append({
                    "cufe": cufe, "numero_factura": None, "proveedor_nombre": None, "total": None,
                    "enviable": False, "motivos_bloqueo": ["No existe una factura con ese CUFE en esta empresa."],
                    "payload": None,
                })
                continue
            armado = siigo_payload.construir_payload(factura, catalogo_taxes)
            motivos = list(armado["motivos_bloqueo"])
            motivo_dup = _motivo_duplicado(conn, factura)
            if motivo_dup:
                motivos.append(motivo_dup)
            enviable = not motivos
            resultado.append({
                "cufe": cufe, "numero_factura": factura["numero_factura"],
                "proveedor_nombre": factura["proveedor_nombre"], "total": factura["total_pagar_xml"],
                "enviable": enviable,
                "motivos_bloqueo": motivos,
                "payload": armado["payload"] if enviable else None,
            })
        return resultado
    finally:
        conn.close()


def confirmar_envio_siigo(slug: str, cufes: list[str]) -> dict:
    """Autentica y envía de verdad a Siigo, una factura a la vez (no hay
    endpoint de lote) -- el único lugar de todo el proyecto donde eso pasa.
    Reconstruye el payload de cada factura en el momento (no confía en uno
    guardado de una previsualización anterior, por si algo cambió entre
    medio) y persiste el resultado -- éxito (`estado_siigo='enviado'`,
    `siigo_id`) o error (`estado_siigo='error'`, `siigo_error` con el
    payload enviado + la respuesta, para poder diagnosticar sin ir a la
    base de datos a mano)."""
    empresa = resolver_empresa(slug)
    conexion = obtener_conexion_siigo(slug)
    if not conexion["configurado"]:
        raise ValueError(
            "Esta empresa todavía no tiene usuario/access_key configurados -- "
            "complétalos primero en el menú 'Conexión Siigo'."
        )
    token = siigo_client.autenticar(conexion["usuario"], conexion["access_key"])
    partner_id = conexion["partner_id"]

    facturas_por_cufe = {f["cufe"]: f for f in listar_facturas(slug)}
    conn = state_store.conectar(empresa["nit"])
    try:
        catalogo_taxes = state_store.listar_catalogo_siigo(conn, "taxes")

        enviadas, con_error = 0, 0
        detalle = []
        nits_asegurados: set[str] = set()  # un solo chequeo/creación por proveedor por lote
        for cufe in cufes:
            factura = facturas_por_cufe.get(cufe)
            if factura is None:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "No existe esa factura en esta empresa."})
                continue

            armado = siigo_payload.construir_payload(factura, catalogo_taxes)
            motivos = list(armado["motivos_bloqueo"])
            motivo_dup = _motivo_duplicado(conn, factura)
            if motivo_dup:
                motivos.append(motivo_dup)
            if motivos:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "; ".join(motivos)})
                continue

            if factura["proveedor_nit"] not in nits_asegurados:
                error_tercero = _asegurar_proveedor_en_siigo(conn, token, partner_id, factura)
                if error_tercero:
                    con_error += 1
                    detalle.append({"cufe": cufe, "ok": False, "error": error_tercero})
                    continue
                nits_asegurados.add(factura["proveedor_nit"])

            try:
                respuesta = siigo_client.crear_purchase(token, partner_id, armado["payload"])
                siigo_id = str(respuesta.get("id", ""))
                state_store.registrar_resultado_envio_siigo(conn, cufe, "enviado", siigo_id=siigo_id, siigo_error=None)
                # Alimentar el caché de compras causadas con lo recién creado
                # (la respuesta de POST /v1/purchases tiene la misma forma que
                # GET /v1/purchases) -- así la protección antidúplicados ve
                # este envío de inmediato, sin esperar a una re-descarga.
                try:
                    fila_cache = _mapear_compra_siigo(respuesta, factura["proveedor_nombre"])
                    state_store.guardar_compras_siigo(conn, [fila_cache], reemplazar_todo=False)
                except (KeyError, TypeError, ValueError):
                    pass  # el caché es una comodidad -- un formato inesperado no debe tumbar un envío ya exitoso
                enviadas += 1
                detalle.append({"cufe": cufe, "ok": True, "siigo_id": siigo_id})
            except siigo_client.SiigoError as e:
                error_json = json.dumps(
                    {"error": str(e), "payload_enviado": armado["payload"]}, ensure_ascii=False,
                )
                state_store.registrar_resultado_envio_siigo(conn, cufe, "error", siigo_id=None, siigo_error=error_json)
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": str(e)})
        return {"enviadas": enviadas, "con_error": con_error, "detalle": detalle}
    finally:
        conn.close()


def previsualizar_exportacion_contai(slug: str, cufes: list[str]) -> list[dict]:
    """Arma, SIN escribir nada, las filas de asiento exactas que se
    exportarían a Contai por cada cufe pedido -- mismo espíritu que
    `previsualizar_envio_siigo` (CLAUDE.md regla 3: nunca se genera/entrega
    un archivo sin que el usuario haya visto antes exactamente qué trae).
    También marca si el proveedor de esa factura es un tercero nuevo (no
    está todavía en el maestro de Contai cacheado)."""
    empresa = resolver_empresa(slug)
    facturas_por_cufe = {f["cufe"]: f for f in listar_facturas(slug)}
    config_contai = obtener_config_contai(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        tipos_cuenta = {c["codigo"]: c["tipo_cuenta"] for c in state_store.listar_plan_cuentas_contai(conn)}
        resultado = []
        for cufe in cufes:
            factura = facturas_por_cufe.get(cufe)
            if factura is None:
                resultado.append({
                    "cufe": cufe, "numero_factura": None, "proveedor_nombre": None, "total": None,
                    "exportable": False, "motivos_bloqueo": ["No existe una factura con ese CUFE en esta empresa."],
                    "filas": None, "tercero_nuevo": False,
                })
                continue
            armado = contai_export.construir_movimientos(factura, config_contai, tipos_cuenta)
            exportable = not armado["motivos_bloqueo"]
            resultado.append({
                "cufe": cufe, "numero_factura": factura["numero_factura"],
                "proveedor_nombre": factura["proveedor_nombre"], "total": factura["total_pagar_xml"],
                "exportable": exportable,
                "motivos_bloqueo": armado["motivos_bloqueo"],
                "filas": armado["filas"],
                "tercero_nuevo": not state_store.existe_tercero_contai(conn, factura["proveedor_nit"]),
            })
        return resultado
    finally:
        conn.close()


def confirmar_exportacion_contai(slug: str, cufes: list[str]) -> dict:
    """Reconstruye cada asiento (no confía en uno guardado de una
    previsualización anterior) y arma CUATRO archivos en memoria, dos
    formatos por cada uno de los dos contenidos (pedido explícito del
    usuario, agosto 2026): movimientos.xlsx/.txt (todas las filas de todas
    las facturas exportables) y terceros.xlsx/.txt (deduplicados por NIT --
    SIEMPRE se genera, con TODOS los proveedores de las facturas que sí
    quedaron exportadas, sean o no nuevos para el maestro de Contai; antes
    solo incluía los que faltaban en `terceros_contai`). El .txt usa el
    mismo separador para ambos archivos (ver contai_export.SEPARADOR_TXT,
    confirmado con el usuario -- punto y coma). Solo marca
    `estado_contai='exportado'` para las facturas que sí quedaron en el
    libro de movimientos -- Contai no tiene API que pueda "rechazar" el
    archivo, así que no hay un estado de error análogo al de Siigo; si algo
    falla acá es un bug de este código, no una respuesta del otro sistema."""
    import openpyxl

    empresa = resolver_empresa(slug)
    facturas_por_cufe = {f["cufe"]: f for f in listar_facturas(slug)}
    config_contai = obtener_config_contai(slug)
    conn = state_store.conectar(empresa["nit"])
    try:
        tipos_cuenta = {c["codigo"]: c["tipo_cuenta"] for c in state_store.listar_plan_cuentas_contai(conn)}

        wb_movimientos = openpyxl.Workbook()
        ws_movimientos = wb_movimientos.active
        ws_movimientos.append(list(contai_export.COLUMNAS_MOVIMIENTO))

        filas_movimiento: list[dict] = []
        terceros: dict[str, dict] = {}
        exportadas, con_error = 0, 0
        detalle = []
        for cufe in cufes:
            factura = facturas_por_cufe.get(cufe)
            if factura is None:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "No existe esa factura en esta empresa."})
                continue

            armado = contai_export.construir_movimientos(factura, config_contai, tipos_cuenta)
            if armado["motivos_bloqueo"]:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "; ".join(armado["motivos_bloqueo"])})
                continue

            for fila in armado["filas"]:
                ws_movimientos.append([fila[col] for col in contai_export.COLUMNAS_MOVIMIENTO])
            filas_movimiento.extend(armado["filas"])

            nit = factura["proveedor_nit"]
            if nit not in terceros:
                fila_origen = conn.execute(
                    "SELECT archivo_origen FROM compras WHERE cufe = ?", (cufe,)
                ).fetchone()
                archivo_origen = fila_origen[0] if fila_origen else None
                tercero = _extraer_tercero_de_origen(archivo_origen, cufe) if archivo_origen else None
                if tercero and tercero.get("nit"):
                    terceros[nit] = contai_export.construir_tercero_nuevo(tercero)

            state_store.registrar_exportacion_contai(conn, cufe, "exportado")
            exportadas += 1
            detalle.append({"cufe": cufe, "ok": True})

        buffer_movimientos = io.BytesIO()
        wb_movimientos.save(buffer_movimientos)
        txt_movimientos = contai_export.filas_a_txt(contai_export.COLUMNAS_MOVIMIENTO, filas_movimiento)

        wb_terceros = openpyxl.Workbook()
        ws_terceros = wb_terceros.active
        ws_terceros.append(list(contai_export.COLUMNAS_TERCERO))
        for fila in terceros.values():
            ws_terceros.append([fila[col] for col in contai_export.COLUMNAS_TERCERO])
        buffer_terceros = io.BytesIO()
        wb_terceros.save(buffer_terceros)
        txt_terceros = contai_export.filas_a_txt(contai_export.COLUMNAS_TERCERO, list(terceros.values()))

        return {
            "exportadas": exportadas, "con_error": con_error, "detalle": detalle,
            "movimientos_xlsx": buffer_movimientos.getvalue(), "movimientos_txt": txt_movimientos,
            "terceros_xlsx": buffer_terceros.getvalue(), "terceros_txt": txt_terceros,
        }
    finally:
        conn.close()


def previsualizar_eliminacion_siigo(slug: str, desde: str, hasta: str, cufes: list[str] | None = None) -> list[dict]:
    """Facturas de esta empresa, en el rango de fechas, que YA están
    enviadas a Siigo -- candidatas para la herramienta de borrado masivo de
    desarrollo/pruebas (ver confirmar_eliminacion_siigo). No toca la red.

    `cufes`, si viene, acota las candidatas a exactamente esas (además del
    rango y de que estén 'enviado') -- así el borrado por período nunca
    incluye facturas que el usuario no marcó explícitamente en la bandeja,
    aunque estén dentro del rango de fechas."""
    if not desde or not hasta:
        raise ValueError("Selecciona un rango de fechas (desde/hasta) antes de borrar por período.")
    facturas = listar_facturas(slug)
    cufes_permitidos = set(cufes) if cufes else None
    candidatas = [
        f for f in facturas
        if f["estado_siigo"] == "enviado" and desde <= f["fecha_emision"] <= hasta
        and (cufes_permitidos is None or f["cufe"] in cufes_permitidos)
    ]
    return [
        {
            "cufe": f["cufe"], "numero_factura": f["numero_factura"], "proveedor_nombre": f["proveedor_nombre"],
            "proveedor_nit": f["proveedor_nit"], "fecha_emision": f["fecha_emision"],
            "total": f["total_pagar_xml"], "siigo_id": f["siigo_id"],
        }
        for f in candidatas
    ]


def _indice_payment_receipts(token: str, partner_id: str, max_paginas: int = 50) -> dict[tuple[str, str], str]:
    """{(nit_proveedor, prefijo+número de factura del proveedor): id del
    recibo} de TODOS los recibos de pago/egreso de la cuenta -- los filtros
    de fecha documentados (`created_start`/`created_end`) NO filtran nada
    contra la API real (confirmado empíricamente: mismo total_results con
    un rango de 1 día que con años completos -- igual al hallazgo ya
    documentado para /v1/purchases), así que se trae todo y se cruza acá.
    `items[].due.{prefix,consecutive}` de un recibo real es exactamente el
    prefijo+número de la factura del PROVEEDOR (confirmado: 'FE'+14040 para
    la factura FE14040 de TORACHE), no un identificador interno de Siigo."""
    indice: dict[tuple[str, str], str] = {}
    page = 1
    while page <= max_paginas:
        resultados, _ = siigo_client.obtener_payment_receipts_pagina(
            token, partner_id, "2000-01-01", "2100-01-01", page, 100,
        )
        if not resultados:
            break
        for r in resultados:
            nit = r.get("supplier", {}).get("identification")
            if not nit:
                continue
            for item in r.get("items", []):
                due = item.get("due") or {}
                ref = f"{due.get('prefix') or ''}{due.get('consecutive') or ''}"
                if ref:
                    indice[(nit, ref)] = r.get("id")
        if len(resultados) < 100:
            break
        page += 1
    return indice


def confirmar_eliminacion_siigo(slug: str, cufes: list[str]) -> dict:
    """Borra en Siigo (primero el recibo de pago automático, después la
    compra) las facturas indicadas, y las deja localmente como 'pendiente'
    otra vez, listas para reenviar -- herramienta de desarrollo/pruebas para
    esta etapa de montaje e implementación, NUNCA parte del flujo normal.
    Requiere que el usuario ya haya confirmado explícitamente cuáles cufes
    borrar (ver previsualizar_eliminacion_siigo) -- acá no se pide ninguna
    confirmación adicional."""
    empresa = resolver_empresa(slug)
    conexion = obtener_conexion_siigo(slug)
    if not conexion["configurado"]:
        raise ValueError(
            "Esta empresa todavía no tiene usuario/access_key configurados -- "
            "complétalos primero en el menú 'Conexión Siigo'."
        )
    token = siigo_client.autenticar(conexion["usuario"], conexion["access_key"])
    partner_id = conexion["partner_id"]

    facturas_por_cufe = {f["cufe"]: f for f in listar_facturas(slug)}
    indice_recibos = _indice_payment_receipts(token, partner_id)

    conn = state_store.conectar(empresa["nit"])
    try:
        eliminadas, con_error = 0, 0
        detalle = []
        for cufe in cufes:
            factura = facturas_por_cufe.get(cufe)
            if factura is None:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "No existe esa factura en esta empresa."})
                continue
            if factura.get("estado_siigo") != "enviado" or not factura.get("siigo_id"):
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": "Esta factura no está marcada como enviada -- no hay nada que borrar en Siigo."})
                continue

            factura_proveedor = _factura_proveedor_de(factura)
            receipt_id = indice_recibos.get((factura["proveedor_nit"], factura_proveedor))

            if receipt_id:
                try:
                    siigo_client.eliminar_payment_receipt(token, partner_id, receipt_id)
                except siigo_client.SiigoError as e:
                    con_error += 1
                    detalle.append({"cufe": cufe, "ok": False, "error": f"No se pudo borrar el recibo de pago ({receipt_id}), no se tocó la compra: {e}"})
                    continue

            try:
                siigo_client.eliminar_purchase(token, partner_id, factura["siigo_id"])
            except siigo_client.SiigoError as e:
                con_error += 1
                detalle.append({"cufe": cufe, "ok": False, "error": f"Recibo borrado pero la compra ({factura['siigo_id']}) no se pudo borrar -- requiere atención manual: {e}"})
                continue

            state_store.registrar_resultado_envio_siigo(conn, cufe, "pendiente", siigo_id=None, siigo_error=None)
            conn.execute("DELETE FROM compras_siigo WHERE siigo_id = ?", (factura["siigo_id"],))
            conn.commit()
            eliminadas += 1
            detalle.append({"cufe": cufe, "ok": True, "recibo_borrado": receipt_id is not None})
        return {"eliminadas": eliminadas, "con_error": con_error, "detalle": detalle}
    finally:
        conn.close()
